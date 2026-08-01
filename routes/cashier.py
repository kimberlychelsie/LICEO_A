from flask import Blueprint, render_template, request, redirect, session, flash, url_for, jsonify, Response
import csv
import io
from db import get_db_connection, is_branch_active
from datetime import datetime, date
from decimal import Decimal
import secrets
import psycopg2.extras
import pandas as pd
from openpyxl.styles import Font, Fill, Alignment, PatternFill, Border, Side
from utils.uniform_pricing import DEFAULT_SIZE_PRICE_STEP, parse_size_list, size_price_map, price_for_size

cashier_bp = Blueprint("cashier", __name__)


def _get_manila_now():
    import pytz
    return datetime.now(pytz.timezone("Asia/Manila"))

def _get_manila_today():
    return _get_manila_now().date()

def _to_manila_naive(dt_value):
    if not dt_value:
        return None
    import pytz
    ph_tz = pytz.timezone("Asia/Manila")
    # If the datetime is naive, assume it's UTC (Postgres default)
    if getattr(dt_value, "tzinfo", None) is None:
        dt_value = pytz.utc.localize(dt_value)
    return dt_value.astimezone(ph_tz).replace(tzinfo=None)

def generate_receipt_number():
    """Generate unique receipt number: OR-YYYYMMDD-XXXXX"""
    today = _get_manila_now().strftime("%Y%m%d")
    random_part = secrets.token_hex(3).upper()  # 6 character hex
    return f"OR-{today}-{random_part}"


def _require_cashier():
    return session.get("role") == "cashier"


def _get_active_year_id(cursor, branch_id):
    cursor.execute("""
        SELECT year_id FROM school_years
        WHERE branch_id = %s AND is_active = TRUE
        LIMIT 1
    """, (branch_id,))
    res = cursor.fetchone()
    return res["year_id"] if res else None


@cashier_bp.route("/cashier")
def dashboard():
    if not _require_cashier():
        return redirect("/")
    if not session.get("branch_id"):
        flash("No branch assigned. Please contact admin.", "error")
        return redirect(url_for("auth.login"))

    db = get_db_connection()
    cursor = db.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

    try:
        active_year_id = _get_active_year_id(cursor, session.get("branch_id"))


        cursor.execute("""
            SELECT
              COUNT(*) AS payment_count,
              COALESCE(SUM(amount), 0) AS total_collected
            FROM payments
            WHERE payment_date::date = %s
              AND branch_id = %s
              AND received_by = %s
              AND year_id = %s
        """, (_get_manila_today(), session.get("branch_id"), session.get("user_id"), active_year_id))
        today_summary = cursor.fetchone() or {"payment_count": 0, "total_collected": 0}

        cursor.execute("""
            SELECT COUNT(*) AS pending_count
            FROM billing b
            JOIN enrollments e ON b.enrollment_id = e.enrollment_id
            WHERE e.branch_id = %s AND b.year_id = %s
              AND b.status IN ('pending', 'partial')
        """, (session.get("branch_id"), active_year_id))
        pending_info = cursor.fetchone() or {"pending_count": 0}

        return render_template(
            "cashier_dashboard.html",
            today_summary=today_summary,
            pending_count=pending_info["pending_count"]
        )
    finally:
        cursor.close()
        db.close()


@cashier_bp.route("/cashier/billing-registry")
def billing_registry():
    if not _require_cashier():
        return redirect("/")
    if not session.get("branch_id"):
        flash("No branch assigned. Please contact admin.", "error")
        return redirect(url_for("auth.login"))

    db = get_db_connection()
    cursor = db.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

    try:
        active_year_id = _get_active_year_id(cursor, session.get("branch_id"))

        # Filters
        status_filter = request.args.get("status_filter", "")
        grade_filter_raw = request.args.get("grade_filter")
        search_q      = request.args.get("q", "").strip()

        # Grade levels for filter dropdown
        cursor.execute("""
            SELECT DISTINCT grade_level FROM enrollments
            WHERE branch_id = %s AND year_id = %s AND status IN ('approved', 'enrolled')
            ORDER BY grade_level
        """, (session.get("branch_id"), active_year_id))
        grade_levels = [r["grade_level"] for r in cursor.fetchall()]

        if grade_filter_raw is None:
            if "Nursery" in grade_levels:
                grade_filter = "Nursery"
            elif any("nursery" in g.lower() for g in grade_levels):
                grade_filter = next(g for g in grade_levels if "nursery" in g.lower())
            elif grade_levels:
                grade_filter = grade_levels[0]
            else:
                grade_filter = "all"
        else:
            grade_filter = grade_filter_raw.strip()

        sql_grade_filter = "" if grade_filter.lower() == "all" else grade_filter

        query = """
            SELECT e.*, b.bill_id, b.balance, b.status AS bill_status,
                   b.total_amount, b.amount_paid,
                   sa.username
            FROM enrollments e
            LEFT JOIN billing b
              ON e.enrollment_id = b.enrollment_id
            LEFT JOIN student_accounts sa
              ON e.enrollment_id = sa.enrollment_id
            WHERE e.branch_id = %s AND e.year_id = %s AND e.status IN ('approved', 'enrolled')
        """
        params = [session.get("branch_id"), active_year_id]

        if sql_grade_filter:
            query += " AND e.grade_level = %s"
            params.append(sql_grade_filter)

        if search_q:
            query += " AND (CONCAT_WS(' ', e.student_first_name ,e.student_middle_name, e.student_last_name) ILIKE %s OR CAST(e.branch_enrollment_no AS TEXT) ILIKE %s)"
            params += [f"%{search_q}%", f"%{search_q}%"]

        if status_filter == "no_bill":
            query += " AND b.bill_id IS NULL"
        elif status_filter == "paid":
            query += " AND b.status = 'paid'"
        elif status_filter == "partial":
            query += " AND b.status = 'partial'"
        elif status_filter == "pending":
            query += " AND b.status = 'pending'"

        query += """
            ORDER BY
              CASE
                WHEN b.bill_id IS NULL THEN 0
                WHEN b.status = 'pending' THEN 1
                WHEN b.status = 'partial' THEN 2
                ELSE 3
              END,
              e.student_last_name ASC,
              e.student_first_name ASC,
              e.student_middle_name ASC
        """

        cursor.execute(query, params)
        enrollments = cursor.fetchall()
        for e in enrollments:
            e["student_name"] = " ".join(filter(None, [
                e.get("student_first_name"),
                e.get("student_middle_name"),
                e.get("student_last_name"),
            ]))

        # Grade levels for filter dropdown
        cursor.execute("""
            SELECT DISTINCT grade_level FROM enrollments
            WHERE branch_id = %s AND year_id = %s AND status IN ('approved', 'enrolled')
            ORDER BY grade_level
        """, (session.get("branch_id"), active_year_id))
        grade_levels = [r["grade_level"] for r in cursor.fetchall()]

        # Summary stats
        paid_count    = sum(1 for e in enrollments if e["bill_status"] == "paid")
        partial_count = sum(1 for e in enrollments if e["bill_status"] == "partial")
        pending_count = sum(1 for e in enrollments if e["bill_status"] == "pending")
        no_bill_count = sum(1 for e in enrollments if not e["bill_id"])

        is_branch_active_status = is_branch_active(session.get("branch_id"))

        return render_template(
            "cashier_billing_registry.html",
            enrollments=enrollments,
            grade_levels=grade_levels,
            status_filter=status_filter,
            grade_filter=grade_filter,
            search_q=search_q,
            paid_count=paid_count,
            partial_count=partial_count,
            pending_count=pending_count,
            no_bill_count=no_bill_count,
            is_branch_active_status=is_branch_active_status,
        )
    finally:
        cursor.close()
        db.close()


@cashier_bp.route("/cashier/create-bill/<int:enrollment_id>", methods=["GET", "POST"])
def create_bill(enrollment_id):
    if not _require_cashier():
        return redirect("/")

    if not is_branch_active(session.get("branch_id")):
        flash("This branch is currently deactivated. New billing records are not allowed.", "error")
        return redirect(url_for("cashier.dashboard"))

    db = get_db_connection()
    cursor = db.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

    try:
        cursor.execute("""
            SELECT e.*, b.branch_name
            FROM enrollments e
            JOIN branches b ON e.branch_id = b.branch_id
            WHERE e.enrollment_id = %s AND e.branch_id = %s
        """, (enrollment_id, session.get("branch_id")))
        enrollment = cursor.fetchone()

        if not enrollment:
            flash("Enrollment not found", "error")
            return redirect("/cashier")

        cursor.execute("""SELECT * FROM billing WHERE enrollment_id = %s""", (enrollment_id,))
        existing_bill = cursor.fetchone()

        if existing_bill:
            flash("Bill already exists for this enrollment", "warning")
            return redirect(url_for("cashier.view_bill", bill_id=existing_bill["bill_id"]))

        # Fetching reservations (both student and parent initiated)
        cursor.execute("""
            SELECT
                ii.category, ii.item_name, ri.qty,
                COALESCE(NULLIF(TRIM(ri.size_label), ''), ii.size_label) as size_label,
                ri.line_total
            FROM reservation_items ri
            JOIN reservations r ON r.reservation_id = ri.reservation_id
            JOIN inventory_items ii ON ri.item_id = ii.item_id
            WHERE r.branch_id = %s AND UPPER(r.status) NOT IN ('CANCELLED', 'REJECTED')
              AND (
                r.enrollment_id = %s
                OR
                r.student_user_id IN (
                    SELECT u.user_id
                    FROM users u
                    JOIN student_accounts sa ON u.username = sa.username
                    WHERE sa.enrollment_id = %s
                )
              )
        """, (session.get("branch_id"), enrollment_id, enrollment_id))
        reservation_items = cursor.fetchall()
        
        books = []
        uniforms = []
        books_total = Decimal("0")
        uniform_total = Decimal("0")

        for item in reservation_items:
            cat = (item["category"] or "").upper()
            if cat == "BOOK":
                books.append({"book_name": item["item_name"], "quantity": item["qty"]})
                books_total += Decimal(str(item["line_total"] or "0"))
            elif cat == "UNIFORM":
                uniforms.append({"uniform_type": item["item_name"], "size": item["size_label"] or "N/A", "quantity": item["qty"]})
                uniform_total += Decimal(str(item["line_total"] or "0"))

        # Legacy items (no price info directly attached, keeping for display)
        cursor.execute("SELECT * FROM enrollment_books WHERE enrollment_id = %s", (enrollment_id,))
        for b in cursor.fetchall():
            books.append({"book_name": b["book_name"] + " (Legacy)", "quantity": b["quantity"]})

        cursor.execute("SELECT * FROM enrollment_uniforms WHERE enrollment_id = %s", (enrollment_id,))
        for u in cursor.fetchall():
            uniforms.append({"uniform_type": u["uniform_type"] + " (Legacy)", "size": u["size"], "quantity": u["quantity"]})

        if request.method == "POST":
            tuition_fee = Decimal(request.form.get("tuition_fee", "0") or "0")
            if tuition_fee > 20000:
                tuition_fee = Decimal("20000")

            books_fee = Decimal(request.form.get("books_fee", "0") or "0")
            uniform_fee = Decimal(request.form.get("uniform_fee", "0") or "0")
            other_fees = Decimal(request.form.get("other_fees", "0") or "0")

            total_amount = tuition_fee + books_fee + uniform_fee + other_fees

            try:
                cursor.execute("""
                    INSERT INTO billing
                      (enrollment_id, branch_id, year_id, tuition_fee, books_fee, uniform_fee, other_fees,
                       total_amount, amount_paid, balance, status, created_by)
                    VALUES
                      (%s, %s, %s, %s, %s, %s, %s,
                       %s, %s, %s, 'pending', %s)
                    RETURNING bill_id
                """, (
                    enrollment_id,
                    session.get("branch_id"),
                    enrollment["year_id"],
                    tuition_fee,
                    books_fee,
                    uniform_fee,
                    other_fees,
                    total_amount,
                    0,
                    total_amount,
                    session.get("user_id")
                ))
                bill_id = cursor.fetchone()["bill_id"]
                db.commit()

                flash(f"Bill created successfully! Total: ₱{total_amount:,.2f}", "success")
                return redirect(url_for("cashier.view_bill", bill_id=bill_id))

            except Exception as e:
                db.rollback()
                flash(f"Failed to create bill: {str(e)}", "error")

        return render_template(
            "cashier_create_bill.html",
            enrollment=enrollment,
            books=books,
            uniforms=uniforms,
            books_total=books_total,
            uniform_total=uniform_total
        )
    finally:
        cursor.close()
        db.close()


@cashier_bp.route("/cashier/bill/<int:bill_id>")
def view_bill(bill_id):
    if not _require_cashier():
        return redirect("/")

    db = get_db_connection()
    cursor = db.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

    try:
        cursor.execute("""
            SELECT
    b.*,
    e.student_first_name,
    e.student_middle_name,
    e.student_last_name,
    e.guardian_first_name,
    e.guardian_middle_name,
    e.guardian_last_name,
    e.grade_level,
    e.branch_enrollment_no,
    br.branch_name,
    u.username AS created_by_name
            FROM billing b
            JOIN enrollments e ON b.enrollment_id = e.enrollment_id
            JOIN branches br ON e.branch_id = br.branch_id
            JOIN users u ON b.created_by = u.user_id
            WHERE b.bill_id = %s AND e.branch_id = %s
        """, (bill_id, session.get("branch_id")))
        bill = cursor.fetchone()

        if not bill:
            flash("Bill not found", "error")
            return redirect("/cashier")

        bill["student_name"] = " ".join(filter(None, [
            bill.get("student_first_name"),
            bill.get("student_middle_name"),
            bill.get("student_last_name"),
        ]))

        bill["guardian_name"] = " ".join(filter(None, [
            bill.get("guardian_first_name"),
            bill.get("guardian_middle_name"),
            bill.get("guardian_last_name"),
        ]))

        cursor.execute("""
            SELECT p.*, COALESCE(NULLIF(TRIM(u.full_name), ''), u.username) AS received_by_name
            FROM payments p
            JOIN users u ON p.received_by = u.user_id
            WHERE p.bill_id = %s
            ORDER BY p.payment_date DESC
        """, (bill_id,))
        payments = cursor.fetchall()

        cursor.execute("SELECT SUM(amount) AS sum_amount FROM payments WHERE bill_id = %s AND target_type = 'tuition'", (bill_id,))
        tuition_paid = cursor.fetchone()["sum_amount"] or Decimal(0)
        
        cursor.execute("SELECT SUM(amount) AS sum_amount FROM payments WHERE bill_id = %s AND target_type = 'other'", (bill_id,))
        other_paid = cursor.fetchone()["sum_amount"] or Decimal(0)

        cursor.execute("SELECT SUM(amount) AS sum_amount FROM payments WHERE bill_id = %s AND target_type = 'general'", (bill_id,))
        general_paid = cursor.fetchone()["sum_amount"] or Decimal(0)

        tuition_balance = bill["tuition_fee"] - tuition_paid
        if general_paid > 0:
            deduct = min(tuition_balance, general_paid)
            tuition_balance -= deduct
            general_paid -= deduct
            
        other_balance = bill["other_fees"] - other_paid
        if general_paid > 0:
            deduct = min(other_balance, general_paid)
            other_balance -= deduct
            general_paid -= deduct

        bill["tuition_balance"] = max(Decimal(0), tuition_balance)
        bill["other_balance"] = max(Decimal(0), other_balance)

        # Fetch grouped reservations for breakdown (books/materials only — uniforms via UO)
        cursor.execute("""
            SELECT
                r.reservation_id, 
                STRING_AGG(ii.item_name || CASE WHEN ri.qty > 1 THEN ' (x' || ri.qty || ')' ELSE '' END, ', ') AS item_names,
                SUM(ri.line_total) AS reservation_total,
                r.status,
                COALESCE((SELECT SUM(amount) FROM payments WHERE target_type='reservation' AND target_id=r.reservation_id), 0) AS paid_amount
            FROM reservation_items ri
            JOIN reservations r ON r.reservation_id = ri.reservation_id
            JOIN inventory_items ii ON ri.item_id = ii.item_id
            WHERE (r.enrollment_id = %s OR r.student_user_id IN (
                SELECT u.user_id FROM student_accounts sa 
                JOIN users u ON sa.username = u.username 
                WHERE sa.enrollment_id = %s
            )) 
            AND UPPER(r.status) NOT IN ('CANCELLED', 'REJECTED')
            AND UPPER(COALESCE(ii.category, '')) <> 'UNIFORM'
            GROUP BY r.reservation_id, r.status
            HAVING COUNT(*) FILTER (WHERE UPPER(COALESCE(ii.category, '')) <> 'UNIFORM') > 0
            ORDER BY r.reservation_id ASC
        """, (bill["enrollment_id"], bill["enrollment_id"]))
        reservation_details = cursor.fetchall()
        
        for res in reservation_details:
            res_bal = res["reservation_total"] - res["paid_amount"]
            if general_paid > 0:
                deduct = min(res_bal, general_paid)
                res_bal -= deduct
                general_paid -= deduct
            res["balance"] = max(Decimal(0), res_bal)

        # Uniform pre-orders linked to this bill (itemized like reservations)
        cursor.execute("""
            SELECT
                uo.order_id,
                uo.order_number,
                uo.order_status,
                uo.payment_status,
                uo.total_amount AS order_total,
                STRING_AGG(
                    uoi.item_name
                    || CASE WHEN uoi.size_label IS NOT NULL AND uoi.size_label <> ''
                            THEN ' (' || uoi.size_label || ')' ELSE '' END
                    || CASE WHEN uoi.quantity > 1 THEN ' x' || uoi.quantity ELSE '' END,
                    ', ' ORDER BY uoi.item_name
                ) AS item_names,
                COALESCE((
                    SELECT SUM(amount) FROM payments
                    WHERE target_type = 'uniform_order' AND target_id = uo.order_id
                ), 0) AS paid_amount
            FROM uniform_orders uo
            JOIN uniform_order_items uoi ON uoi.order_id = uo.order_id
            WHERE uo.bill_id = %s
               OR (uo.enrollment_id = %s AND uo.order_status IN ('Ready for Claim', 'Claimed'))
            GROUP BY uo.order_id, uo.order_number, uo.order_status, uo.payment_status, uo.total_amount
            ORDER BY uo.order_id ASC
        """, (bill_id, bill["enrollment_id"]))
        uniform_order_details = cursor.fetchall() or []

        for uo in uniform_order_details:
            uo_bal = Decimal(str(uo["order_total"] or 0)) - Decimal(str(uo["paid_amount"] or 0))
            if general_paid > 0:
                deduct = min(uo_bal, general_paid)
                uo_bal -= deduct
                general_paid -= deduct
            uo["balance"] = max(Decimal(0), uo_bal)

            # Query itemized set piece breakdown for expanding UI
            cursor.execute("""
                SELECT uoi.item_name, uoi.size_label, uoi.unit_price, uoi.quantity, uoi.line_total, uoi.inventory_item_id
                FROM uniform_order_items uoi
                WHERE uoi.order_id = %s
                ORDER BY uoi.item_name ASC
            """, (uo["order_id"],))
            uoi_items = cursor.fetchall() or []

            pieces_list = []
            for item in uoi_items:
                inv_id = item.get("inventory_item_id")
                size_lbl = item.get("size_label") or ""
                if inv_id:
                    cursor.execute("""
                        SELECT item_name, price, size_label, COALESCE(size_price_step, 20) AS size_price_step
                        FROM inventory_items
                        WHERE parent_item_id = %s AND is_active = TRUE
                        ORDER BY item_name ASC
                    """, (inv_id,))
                    child_pieces = cursor.fetchall()
                    if child_pieces:
                        pieces_subtotal = Decimal(0)
                        for cp in child_pieces:
                            cp_price = price_for_size(cp["price"], size_lbl, cp["size_label"], 0)
                            cp_line_total = Decimal(str(cp_price)) * Decimal(str(item["quantity"]))
                            pieces_subtotal += cp_line_total
                            pieces_list.append({
                                "item_name": cp["item_name"],
                                "size_label": size_lbl,
                                "unit_price": cp_price,
                                "quantity": item["quantity"],
                                "line_total": cp_line_total
                            })
                        item_total = Decimal(str(item["line_total"] or 0))
                        size_fee = item_total - pieces_subtotal
                        if size_fee > Decimal(0):
                            qty = Decimal(str(item["quantity"] or 1))
                            pieces_list.append({
                                "item_name": f"+ Size Fee ({size_lbl})" if size_lbl else "+ Size Fee",
                                "size_label": "",
                                "unit_price": size_fee / qty if qty else size_fee,
                                "quantity": item["quantity"],
                                "line_total": size_fee
                            })
                    else:
                        pieces_list.append({
                            "item_name": item["item_name"],
                            "size_label": size_lbl,
                            "unit_price": Decimal(str(item["unit_price"] or 0)),
                            "quantity": item["quantity"],
                            "line_total": Decimal(str(item["line_total"] or 0))
                        })
                else:
                    pieces_list.append({
                        "item_name": item["item_name"],
                        "size_label": size_lbl,
                        "unit_price": Decimal(str(item["unit_price"] or 0)),
                        "quantity": item["quantity"],
                        "line_total": Decimal(str(item["line_total"] or 0))
                    })
            uo["pieces"] = pieces_list

        # Recalculate true Total Fee, Paid Amount, and Balance across entire bill
        books_sum = sum(Decimal(str(r["reservation_total"] or 0)) for r in reservation_details)
        uniforms_sum = sum(Decimal(str(u["order_total"] or 0)) for u in uniform_order_details)

        calc_tuition = Decimal(str(bill.get("tuition_fee") or 0))
        calc_other = Decimal(str(bill.get("other_fees") or 0))
        calc_total = calc_tuition + calc_other + books_sum + uniforms_sum

        cursor.execute("SELECT COALESCE(SUM(amount), 0) AS total_paid FROM payments WHERE bill_id = %s", (bill_id,))
        calc_paid = cursor.fetchone()["total_paid"] or Decimal(0)
        calc_balance = max(Decimal(0), calc_total - calc_paid)

        if calc_balance <= 0 and calc_total > 0:
            calc_status = "paid"
        elif calc_paid > 0:
            calc_status = "partial"
        else:
            calc_status = "pending"

        bill["books_fee"] = books_sum
        bill["uniform_fee"] = uniforms_sum
        bill["total_amount"] = calc_total
        bill["amount_paid"] = calc_paid
        bill["balance"] = calc_balance
        bill["status"] = calc_status

        # Persist synced totals to DB so Billing Registry & Student/Parent portal match 100%
        cursor.execute("""
            UPDATE billing
            SET books_fee = %s,
                uniform_fee = %s,
                total_amount = %s,
                amount_paid = %s,
                balance = %s,
                status = %s
            WHERE bill_id = %s
        """, (books_sum, uniforms_sum, calc_total, calc_paid, calc_balance, calc_status, bill_id))
        db.commit()

        # Fetch branch admin name (acting as Principal)
        branch_admin_name = _fetch_branch_admin_name(cursor, session.get("branch_id"))

        # Construct payment itemized breakdown list in Python for safety
        for p in payments:
            p_items = []
            amt_val = float(p["amount"] or 0)
            if p["target_type"] == "tuition":
                p_items.append({"desc": "Tuition & Mandatory Fees", "amount": amt_val})
            elif p["target_type"] == "other":
                p_items.append({"desc": "Other Fees", "amount": amt_val})
            elif p["target_type"] == "reservation":
                desc_str = "Reservation (Books/Materials)"
                for res in reservation_details:
                    if res["reservation_id"] == p["target_id"]:
                        desc_str = res["item_names"]
                        break
                p_items.append({"desc": desc_str, "amount": amt_val})
            elif p["target_type"] == "uniform_order":
                desc_str = "Uniform Pre-Order"
                for uo in uniform_order_details:
                    if uo["order_id"] == p["target_id"]:
                        desc_str = uo["item_names"]
                        break
                p_items.append({"desc": desc_str, "amount": amt_val})
            else:
                p_items.append({"desc": p["notes"] or "General Billing Payment", "amount": amt_val})
            p["breakdown_items"] = p_items

        return render_template(
            "cashier_view_bill.html",
            bill=bill,
            payments=payments,
            reservation_details=reservation_details,
            uniform_order_details=uniform_order_details,
            branch_admin_name=branch_admin_name,
        )
    finally:
        cursor.close()
        db.close()


@cashier_bp.route("/cashier/process-payment/<int:bill_id>", methods=["GET", "POST"])
def process_payment(bill_id):
    if not _require_cashier():
        return redirect("/")

    if not is_branch_active(session.get("branch_id")):
        flash("This branch is currently deactivated. New payments are not allowed.", "error")
        return redirect(url_for("cashier.view_bill", bill_id=bill_id))

    db = get_db_connection()
    cursor = db.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

    try:
        cursor.execute("""
            SELECT
    b.*,
    e.student_first_name,
    e.student_middle_name,
    e.student_last_name,
    e.grade_level,
                       e.branch_enrollment_no
            FROM billing b
            JOIN enrollments e ON b.enrollment_id = e.enrollment_id
            WHERE b.bill_id = %s AND e.branch_id = %s
        """, (bill_id, session.get("branch_id")))
        bill = cursor.fetchone()

        if not bill:
            flash("Bill not found", "error")
            return redirect("/cashier")
        bill["student_name"] = " ".join(filter(None, [
            bill.get("student_first_name"),
            bill.get("student_middle_name"),
            bill.get("student_last_name"),
        ]))

        if bill["status"] == "paid":
            flash("This bill is already fully paid", "info")
            return redirect(url_for("cashier.view_bill", bill_id=bill_id))

        if request.method == "POST":
            amount = Decimal(request.form.get("amount", "0") or "0")
            payment_method = request.form.get("payment_method", "cash")
            notes = request.form.get("notes", "")
            target_type = request.form.get("target_type", "general")
            target_id_str = request.form.get("target_id", "")
            target_id = int(target_id_str) if target_id_str.strip() else None

            if amount <= 0:
                flash("Payment amount must be greater than zero", "error")
            elif amount > Decimal(str(bill["balance"])):
                flash(
                    f"Payment amount (₱{amount:,.2f}) exceeds balance (₱{Decimal(str(bill['balance'])):,.2f})",
                    "error"
                )
            else:
                try:
                    receipt_number = generate_receipt_number()

                    cursor.execute("""
                        INSERT INTO payments
                          (bill_id, enrollment_id, branch_id, year_id, amount, payment_method,
                           receipt_number, notes, received_by, target_type, target_id)
                        VALUES
                          (%s, %s, %s, %s, %s, %s,
                           %s, %s, %s, %s, %s)
                        RETURNING payment_id
                    """, (
                        bill_id,
                        bill["enrollment_id"],
                        session.get("branch_id"),
                        bill["year_id"],
                        amount,
                        payment_method,
                        receipt_number,
                        notes,
                        session.get("user_id"),
                        target_type,
                        target_id
                    ))
                    payment_id = cursor.fetchone()["payment_id"]

                    amount_paid_now = Decimal(str(bill.get("amount_paid", 0)))
                    total_amount = Decimal(str(bill.get("total_amount", 0)))

                    new_amount_paid = amount_paid_now + amount
                    new_balance = total_amount - new_amount_paid
                    if new_balance < 0:
                        new_balance = Decimal("0")

                    if target_type == 'reservation' and target_id:
                        cursor.execute("SELECT SUM(amount) AS sum_paid FROM payments WHERE target_type='reservation' AND target_id=%s", (target_id,))
                        r_paid = cursor.fetchone()["sum_paid"] or Decimal(0)
                        
                        cursor.execute("SELECT SUM(line_total) AS sum_total FROM reservation_items WHERE reservation_id=%s", (target_id,))
                        r_total = cursor.fetchone()["sum_total"] or Decimal(0)
                        
                        if r_paid >= r_total:
                            cursor.execute("""
                                UPDATE reservations
                                SET status = 'PAID'
                                WHERE reservation_id = %s
                            """, (target_id,))

                    if target_type == 'uniform_order' and target_id:
                        cursor.execute(
                            "SELECT SUM(amount) AS sum_paid FROM payments WHERE target_type='uniform_order' AND target_id=%s",
                            (target_id,),
                        )
                        u_paid = cursor.fetchone()["sum_paid"] or Decimal(0)
                        cursor.execute(
                            "SELECT total_amount FROM uniform_orders WHERE order_id=%s",
                            (target_id,),
                        )
                        u_row = cursor.fetchone()
                        u_total = Decimal(str((u_row or {}).get("total_amount") or 0))
                        if u_paid >= u_total and u_total > 0:
                            cursor.execute("""
                                UPDATE uniform_orders
                                SET payment_status = 'Paid', updated_at = NOW()
                                WHERE order_id = %s
                            """, (target_id,))

                    if new_balance == 0:
                        new_status = "paid"
                        # Auto-mark active book/material reservations only — never revive CANCELLED
                        # and never touch uniform-only rows (those live in uniform_orders).
                        cursor.execute("""
                            UPDATE reservations r
                            SET status = 'PAID'
                            WHERE (
                                r.enrollment_id = %s OR r.student_user_id IN (
                                    SELECT u.user_id FROM student_accounts sa
                                    JOIN users u ON sa.username = u.username
                                    WHERE sa.enrollment_id = %s
                                )
                            )
                            AND UPPER(COALESCE(r.status, '')) NOT IN ('CANCELLED', 'REJECTED')
                            AND EXISTS (
                                SELECT 1 FROM reservation_items ri
                                JOIN inventory_items ii ON ii.item_id = ri.item_id
                                WHERE ri.reservation_id = r.reservation_id
                                  AND UPPER(COALESCE(ii.category, '')) <> 'UNIFORM'
                            )
                        """, (bill["enrollment_id"], bill["enrollment_id"]))
                        cursor.execute("""
                            UPDATE uniform_orders
                            SET payment_status = 'Paid', updated_at = NOW()
                            WHERE bill_id = %s OR enrollment_id = %s
                        """, (bill_id, bill["enrollment_id"]))
                    else:
                        new_status = "partial"

                    cursor.execute("""
                        UPDATE billing
                        SET amount_paid = %s, balance = %s, status = %s
                        WHERE bill_id = %s
                    """, (new_amount_paid, new_balance, new_status, bill_id))

                    db.commit()

                    flash(f"Payment recorded successfully! Receipt: {receipt_number}", "success")
                    return redirect(url_for("cashier.print_receipt", payment_id=payment_id))

                except Exception as e:
                    db.rollback()
                    flash(f"Failed to process payment: {str(e)}", "error")

        target_type = request.args.get("target_type", "general")
        target_id = request.args.get("target_id", "")
        target_amount = request.args.get("target_amount", "")
        target_name = request.args.get("target_name", "")

        return render_template("cashier_process_payment.html", bill=bill, target_type=target_type, target_id=target_id, target_amount=target_amount, target_name=target_name)
    finally:
        cursor.close()
        db.close()


@cashier_bp.route("/cashier/receipt/<int:payment_id>")
def print_receipt(payment_id):
    if not _require_cashier():
        return redirect("/")

    db = get_db_connection()
    cursor = db.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

    try:
        cursor.execute("""
            SELECT
    p.*,
    e.student_first_name,
    e.student_middle_name,
    e.student_last_name,
    e.guardian_first_name,
    e.guardian_middle_name,
    e.guardian_last_name,
    e.grade_level,
    e.branch_enrollment_no,
    b.total_amount,
    b.amount_paid,
    b.balance,
    br.branch_name,
    br.location,
    COALESCE(NULLIF(TRIM(u.full_name), ''), u.username) AS received_by_name
            FROM payments p
            JOIN billing b ON p.bill_id = b.bill_id
            JOIN enrollments e ON p.enrollment_id = e.enrollment_id
            JOIN branches br ON e.branch_id = br.branch_id
            JOIN users u ON p.received_by = u.user_id
            WHERE p.payment_id = %s
        """, (payment_id,))
        payment = cursor.fetchone()

        if not payment:
            flash("Receipt not found", "error")
            return redirect("/cashier")
        payment["student_name"] = " ".join(filter(None, [
            payment.get("student_first_name"),
            payment.get("student_middle_name"),
            payment.get("student_last_name"),
        ]))

        payment["guardian_name"] = " ".join(filter(None, [
            payment.get("guardian_first_name"),
            payment.get("guardian_middle_name"),
            payment.get("guardian_last_name"),
        ]))

        payment["payment_date"] = _to_manila_naive(payment.get("payment_date"))
        return render_template("cashier_receipt.html", payment=payment)
    finally:
        cursor.close()
        db.close()


def _get_report_date_range(report_range, report_date):
    from datetime import timedelta
    today = _get_manila_today()
    if report_range == "weekly":
        start = today - timedelta(days=today.weekday())
        end   = start + timedelta(days=6)
    elif report_range == "monthly":
        import calendar
        start = today.replace(day=1)
        last  = calendar.monthrange(today.year, today.month)[1]
        end   = today.replace(day=last)
    elif report_range == "yearly":
        start = today.replace(month=1, day=1)
        end   = today.replace(month=12, day=31)
    else:
        try:
            from datetime import date as _date
            d = _date.fromisoformat(report_date) if report_date else today
        except Exception:
            d = today
        start = end = d
    return start, end


def _fetch_report_payments(cursor, branch_id, start_date, end_date):
    cursor.execute("""
        SELECT p.payment_id, p.receipt_number, p.amount, p.payment_method, p.payment_date,
               e.student_first_name, e.student_middle_name, e.student_last_name,
               e.grade_level, e.branch_enrollment_no,
               COALESCE(NULLIF(TRIM(u.full_name),''), u.username) AS received_by_name
        FROM payments p
        JOIN enrollments e ON p.enrollment_id = e.enrollment_id
        JOIN users u ON p.received_by = u.user_id
        WHERE p.payment_date::date BETWEEN %s AND %s
          AND e.branch_id = %s
        ORDER BY p.payment_date DESC
    """, (start_date, end_date, branch_id))
    payments = cursor.fetchall()
    for payment in payments:
        payment["student_name"] = " ".join(filter(None, [
            payment.get("student_first_name"),
            payment.get("student_middle_name"),
            payment.get("student_last_name"),
        ]))
    return payments


def _fetch_report_summary(cursor, branch_id, start_date, end_date):
    cursor.execute("""
        SELECT COUNT(*) AS transaction_count,
               COALESCE(SUM(p.amount), 0) AS total_collected
        FROM payments p
        JOIN enrollments e ON p.enrollment_id = e.enrollment_id
        WHERE p.payment_date::date BETWEEN %s AND %s
          AND e.branch_id = %s
    """, (start_date, end_date, branch_id))
    return cursor.fetchone() or {"transaction_count": 0, "total_collected": 0}


def _fetch_branch_admin_name(cursor, branch_id):
    cursor.execute("""
        SELECT COALESCE(NULLIF(TRIM(full_name), ''), username) AS admin_name
        FROM users WHERE branch_id = %s AND role = 'branch_admin' LIMIT 1
    """, (branch_id,))
    row = cursor.fetchone()
    return row["admin_name"] if row else "Branch Administrator"


@cashier_bp.route("/cashier/reports", methods=["GET", "POST"])
def reports():
    if not _require_cashier():
        return redirect("/")

    report_range = request.form.get("report_range", "today")
    report_date  = request.form.get("report_date", _get_manila_today().strftime("%Y-%m-%d"))
    start_date, end_date = _get_report_date_range(report_range, report_date)

    db = get_db_connection()
    cursor = db.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    try:
        branch_id = session.get("branch_id")
        payments  = _fetch_report_payments(cursor, branch_id, start_date, end_date)
        summary   = _fetch_report_summary(cursor, branch_id, start_date, end_date)
        branch_admin_name = _fetch_branch_admin_name(cursor, branch_id)
        cashier_name = session.get("full_name") or session.get("username") or "Authorized Cashier"

        return render_template(
            "cashier_reports.html",
            payments=payments,
            summary=summary,
            report_date=report_date,
            report_range=report_range,
            date_start=str(start_date),
            date_end=str(end_date),
            cashier_name=cashier_name,
            branch_admin_name=branch_admin_name
        )
    finally:
        cursor.close()
        db.close()


@cashier_bp.route("/cashier/reports/export/excel")
def export_reports_excel():
    if not _require_cashier():
        return redirect("/")

    report_range = request.args.get("range", "today")
    report_date  = request.args.get("date", _get_manila_today().strftime("%Y-%m-%d"))
    start_date, end_date = _get_report_date_range(report_range, report_date)

    db = get_db_connection()
    cursor = db.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    try:
        branch_id = session.get("branch_id")
        payments  = _fetch_report_payments(cursor, branch_id, start_date, end_date)
        summary   = _fetch_report_summary(cursor, branch_id, start_date, end_date)
        branch_admin_name = _fetch_branch_admin_name(cursor, branch_id)
        cashier_name = session.get("full_name") or session.get("username") or "Authorized Cashier"
    finally:
        cursor.close()
        db.close()

    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    import io as _io

    wb = Workbook()
    ws = wb.active
    ws.title = "Financial Report"

    hdr_font  = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    hdr_fill  = PatternFill("solid", fgColor="1E3A8A")
    hdr_align = Alignment(horizontal="center", vertical="center")
    thin_side = Side(style="thin", color="CBD5E1")
    thin_bdr  = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    center_al = Alignment(horizontal="center", vertical="center")
    right_al  = Alignment(horizontal="right", vertical="center")

    ws.merge_cells("A1:G1")
    ws["A1"] = "LICEO DE MAJAYJAY \u2014 FINANCIAL COLLECTION REPORT"
    ws["A1"].font = Font(name="Calibri", bold=True, size=14, color="1E3A8A")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26

    range_label = {
        "today":   f"Today ({start_date})",
        "weekly":  f"This Week ({start_date} to {end_date})",
        "monthly": f"This Month ({start_date} to {end_date})",
        "yearly":  f"This Year ({start_date} to {end_date})",
        "custom":  f"Custom ({start_date})",
    }.get(report_range, str(start_date))

    ws.merge_cells("A2:G2")
    ws["A2"] = f"Period: {range_label}"
    ws["A2"].font = Font(name="Calibri", italic=True, size=10, color="475569")
    ws["A2"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A3:G3")
    ws["A3"] = f"Cashier: {cashier_name}   |   Total Collected: P{float(summary['total_collected']):.2f}   |   Transactions: {summary['transaction_count']}"
    ws["A3"].font = Font(name="Calibri", bold=True, size=10)
    ws["A3"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[3].height = 18

    headers = ["#", "OR Number", "Student Name", "Grade Level", "Payment Method", "Amount", "Date"]
    ws.append([""] * 7)
    ws.row_dimensions[4].height = 6
    ws.append(headers)
    for col_idx in range(1, 8):
        cell = ws.cell(row=5, column=col_idx)
        cell.font      = hdr_font
        cell.fill      = hdr_fill
        cell.alignment = hdr_align
        cell.border    = thin_bdr
    ws.row_dimensions[5].height = 22

    import pytz as _pytz
    ph_tz = _pytz.timezone("Asia/Manila")
    for i, p in enumerate(payments, 1):
        pd_val = p.get("payment_date")
        if pd_val and hasattr(pd_val, "astimezone"):
            pd_val = pd_val.astimezone(ph_tz).strftime("%Y-%m-%d %H:%M")
        row = [
            i, p.get("receipt_number", ""), p.get("student_name", ""),
            p.get("grade_level", ""), str(p.get("payment_method", "")).upper(),
            float(p.get("amount", 0)), str(pd_val) if pd_val else "",
        ]
        ws.append(row)
        data_row = ws.max_row
        for col_idx in range(1, 8):
            cell = ws.cell(row=data_row, column=col_idx)
            cell.border    = thin_bdr
            cell.alignment = right_al if col_idx == 6 else center_al
            cell.font      = Font(name="Calibri", size=10)
        if i % 2 == 0:
            alt_fill = PatternFill("solid", fgColor="F8FAFC")
            for col_idx in range(1, 8):
                ws.cell(row=data_row, column=col_idx).fill = alt_fill

    for i, w in enumerate([5, 14, 28, 14, 18, 14, 20], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    total_row = ws.max_row + 1
    ws.cell(total_row, 5, "TOTAL").font = Font(name="Calibri", bold=True, size=10)
    ws.cell(total_row, 5).alignment     = Alignment(horizontal="right")
    ws.cell(total_row, 6, float(summary["total_collected"])).font = Font(name="Calibri", bold=True, size=10)
    ws.cell(total_row, 6).alignment     = Alignment(horizontal="right")
    ws.cell(total_row, 6).number_format = "#,##0.00"
    for col_idx in range(1, 8):
        ws.cell(total_row, col_idx).border = thin_bdr
        ws.cell(total_row, col_idx).fill   = PatternFill("solid", fgColor="EFF6FF")

    ws.append([""] * 7)
    sig_row = ws.max_row + 2
    ws.cell(sig_row,   1, "Prepared By:").font        = Font(name="Calibri", bold=True)
    ws.cell(sig_row,   5, "Approved By:").font         = Font(name="Calibri", bold=True)
    ws.cell(sig_row+1, 1, cashier_name).font           = Font(name="Calibri", bold=True, underline="single")
    ws.cell(sig_row+1, 5, branch_admin_name).font      = Font(name="Calibri", bold=True, underline="single")
    ws.cell(sig_row+2, 1, "Cashier").font              = Font(name="Calibri", italic=True, color="64748B")
    ws.cell(sig_row+2, 5, "Branch Administrator").font = Font(name="Calibri", italic=True, color="64748B")

    buf = _io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"Financial_Report_{report_range}_{start_date}.xlsx"
    return Response(
        buf.read(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@cashier_bp.route("/cashier/reports/export/pdf")
def export_reports_pdf():
    if not _require_cashier():
        return redirect("/")

    report_range = request.args.get("range", "today")
    report_date  = request.args.get("date", _get_manila_today().strftime("%Y-%m-%d"))
    start_date, end_date = _get_report_date_range(report_range, report_date)

    db = get_db_connection()
    cursor = db.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    try:
        branch_id = session.get("branch_id")
        payments  = _fetch_report_payments(cursor, branch_id, start_date, end_date)
        summary   = _fetch_report_summary(cursor, branch_id, start_date, end_date)
        branch_admin_name = _fetch_branch_admin_name(cursor, branch_id)
        cashier_name = session.get("full_name") or session.get("username") or "Authorized Cashier"
    finally:
        cursor.close()
        db.close()

    from reportlab.lib import colors as rl_colors
    from reportlab.lib.pagesizes import letter
    from reportlab.lib.units import mm
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable
    from reportlab.lib.enums import TA_CENTER
    import io as _io

    buf = _io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=letter,
                            leftMargin=20*mm, rightMargin=20*mm,
                            topMargin=18*mm, bottomMargin=18*mm)

    styles   = getSampleStyleSheet()
    navy     = rl_colors.HexColor("#1E3A8A")
    slate    = rl_colors.HexColor("#475569")
    light_bg = rl_colors.HexColor("#F8FAFC")
    border_c = rl_colors.HexColor("#CBD5E1")

    title_st = ParagraphStyle("title", parent=styles["Normal"],
        fontName="Helvetica-Bold", fontSize=16, textColor=navy,
        alignment=TA_CENTER, spaceAfter=4)
    sub_st = ParagraphStyle("sub", parent=styles["Normal"],
        fontName="Helvetica", fontSize=9, textColor=slate,
        alignment=TA_CENTER, spaceAfter=2)
    label_st = ParagraphStyle("label", parent=styles["Normal"],
        fontName="Helvetica-Bold", fontSize=9, textColor=navy,
        alignment=TA_CENTER, spaceAfter=10)

    range_label = {
        "today":   f"Today: {start_date}",
        "weekly":  f"This Week: {start_date} to {end_date}",
        "monthly": f"This Month: {start_date} to {end_date}",
        "yearly":  f"This Year: {start_date} to {end_date}",
        "custom":  f"Custom Date: {start_date}",
    }.get(report_range, str(start_date))

    story = [
        Paragraph("LICEO DE MAJAYJAY", title_st),
        Paragraph("Financial Collection Report", sub_st),
        Paragraph(range_label, sub_st),
        HRFlowable(width="100%", thickness=2, color=navy, spaceAfter=8),
        Paragraph(
            f"Cashier: <b>{cashier_name}</b> | "
            f"Total: <b>P{float(summary['total_collected']):,.2f}</b> | "
            f"Transactions: <b>{summary['transaction_count']}</b>",
            label_st
        ),
        Spacer(1, 4*mm),
    ]

    col_headers = ["#", "OR Number", "Student Name", "Grade", "Amount", "Date"]
    tbl_data = [col_headers]
    import pytz as _pytz
    ph_tz = _pytz.timezone("Asia/Manila")
    for i, p in enumerate(payments, 1):
        pd_val = p.get("payment_date")
        if pd_val and hasattr(pd_val, "astimezone"):
            pd_val = pd_val.astimezone(ph_tz).strftime("%m/%d/%Y %H:%M")
        tbl_data.append([
            str(i), p.get("receipt_number", ""), p.get("student_name", ""),
            p.get("grade_level", ""), f"P{float(p.get('amount', 0)):,.2f}",
            str(pd_val) if pd_val else "",
        ])
    tbl_data.append(["", "", "", "TOTAL",
                     f"P{float(summary['total_collected']):,.2f}", ""])

    avail_w = letter[0] - 40*mm
    cw = [0.05*avail_w, 0.15*avail_w, 0.32*avail_w,
          0.15*avail_w, 0.16*avail_w, 0.17*avail_w]

    tbl = Table(tbl_data, colWidths=cw, repeatRows=1)
    tbl.setStyle(TableStyle([
        ("BACKGROUND",    (0,0), (-1,0), navy),
        ("TEXTCOLOR",     (0,0), (-1,0), rl_colors.white),
        ("FONTNAME",      (0,0), (-1,0), "Helvetica-Bold"),
        ("FONTSIZE",      (0,0), (-1,0), 8),
        ("ALIGN",         (0,0), (-1,0), "CENTER"),
        ("VALIGN",        (0,0), (-1,-1), "MIDDLE"),
        ("ROWBACKGROUNDS",(0,1), (-1,-2), [light_bg, rl_colors.white]),
        ("FONTNAME",      (0,1), (-1,-1), "Helvetica"),
        ("FONTSIZE",      (0,1), (-1,-1), 8),
        ("ALIGN",         (4,1), (4,-1), "RIGHT"),
        ("BACKGROUND",    (0,-1), (-1,-1), rl_colors.HexColor("#EFF6FF")),
        ("FONTNAME",      (0,-1), (-1,-1), "Helvetica-Bold"),
        ("GRID",          (0,0), (-1,-1), 0.5, border_c),
        ("LINEBELOW",     (0,0), (-1,0), 1.5, navy),
        ("LINEABOVE",     (0,-1), (-1,-1), 1, navy),
        ("TOPPADDING",    (0,0), (-1,-1), 5),
        ("BOTTOMPADDING", (0,0), (-1,-1), 5),
        ("LEFTPADDING",   (0,0), (-1,-1), 5),
        ("RIGHTPADDING",  (0,0), (-1,-1), 5),
    ]))
    story.append(tbl)
    story.append(Spacer(1, 18*mm))

    sig_data = [
        [cashier_name, "", branch_admin_name],
        ["________________________", "", "________________________"],
        ["Cashier", "", "Branch Administrator"],
    ]
    sig_tbl = Table(sig_data, colWidths=[avail_w*0.4, avail_w*0.2, avail_w*0.4])
    sig_tbl.setStyle(TableStyle([
        ("FONTNAME",  (0,0), (-1,0),  "Helvetica-Bold"),
        ("FONTSIZE",  (0,0), (-1,0),  9),
        ("ALIGN",     (0,0), (0,-1),  "CENTER"),
        ("ALIGN",     (2,0), (2,-1),  "CENTER"),
        ("FONTNAME",  (0,2), (-1,2),  "Helvetica-Oblique"),
        ("FONTSIZE",  (0,2), (-1,2),  8),
        ("TEXTCOLOR", (0,2), (-1,2),  slate),
    ]))
    story.append(sig_tbl)

    doc.build(story)
    buf.seek(0)
    filename = f"Financial_Report_{report_range}_{start_date}.pdf"
    return Response(
        buf.read(),
        mimetype="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@cashier_bp.route("/cashier/search", methods=["GET", "POST"])
def search():
    if not _require_cashier():
        return redirect("/")

    results = []
    search_query = ""

    if request.method == "POST":
        search_query = request.form.get("search_query", "").strip()

        if search_query:
            db = get_db_connection()
            cursor = db.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

            try:
                is_numeric = search_query.isdigit()

                cursor.execute("""
                    SELECT e.*, b.bill_id, b.balance, b.status AS bill_status,
                           b.total_amount, b.amount_paid
                    FROM enrollments e
                    LEFT JOIN billing b ON e.enrollment_id = b.enrollment_id
                    WHERE e.branch_id = %s
                      AND (
                        (%s AND (e.branch_enrollment_no = %s))
                        OR (
    CONCAT_WS(
        ' ',
        e.student_first_name,
        e.student_middle_name,
        e.student_last_name
    ) ILIKE %s
)
                      )
                    ORDER BY e.created_at DESC
                """, (
                    session.get("branch_id"),
                    is_numeric,
                    int(search_query) if is_numeric else 0,
                    f"%{search_query}%"
                ))

                results = cursor.fetchall()
                for r in results:
                    r["student_name"] = " ".join(filter(None, [
                        r.get("student_first_name"),
                        r.get("student_middle_name"),
                        r.get("student_last_name"),
                    ]))
            finally:
                cursor.close()
                db.close()

    return render_template("cashier_search.html", results=results, search_query=search_query)


@cashier_bp.route("/cashier/payment-history", methods=["GET"])
def payment_history():
    if not _require_cashier():
        return redirect("/")

    date_from = request.args.get("date_from", "").strip() or _get_manila_today().replace(day=1).strftime("%Y-%m-%d")
    date_to = request.args.get("date_to", "").strip() or _get_manila_today().strftime("%Y-%m-%d")

    db = get_db_connection()
    cursor = db.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

    try:
        cursor.execute("""
            SELECT
              p.payment_id,
              p.receipt_number as reference_number,
              p.amount,
              p.payment_method,
              p.payment_date,
              e.student_first_name,
              e.student_middle_name,
              e.student_last_name,
              e.grade_level,
              u.username AS received_by_name
            FROM payments p
            JOIN enrollments e ON p.enrollment_id = e.enrollment_id
            JOIN users u ON p.received_by = u.user_id
            WHERE p.payment_date::date >= %s AND p.payment_date::date <= %s
              AND e.branch_id = %s
            ORDER BY p.payment_date DESC
        """, (date_from, date_to, session.get("branch_id")))
        payments = cursor.fetchall()

        for payment in payments:
            payment["student_name"] = " ".join(filter(None, [
                payment.get("student_first_name"),
                payment.get("student_middle_name"),
                payment.get("student_last_name"),
            ]))

        cursor.execute("""
            SELECT
              COALESCE(SUM(p.amount), 0) AS total_collected
            FROM payments p
            JOIN enrollments e ON p.enrollment_id = e.enrollment_id
            WHERE p.payment_date::date >= %s AND p.payment_date::date <= %s
              AND e.branch_id = %s
        """, (date_from, date_to, session.get("branch_id")))
        summary = cursor.fetchone() or {"total_collected": 0}

        return render_template(
            "payment_history.html",
            payments=payments,
            total_collected=summary["total_collected"],
            date_from=date_from,
            date_to=date_to
        )
    except Exception as e:
        flash(f"Error loading payment history: {str(e)}", "error")
        return redirect(url_for("cashier.dashboard"))
    finally:
        cursor.close()
        db.close()


# =======================
# CASHIER RESERVATIONS (NO students TABLE)
# =======================

def _normalize_category(raw: str | None) -> str | None:
    if not raw:
        return None
    c = raw.strip().upper()
    # keep it strict to what you support in UI
    if c in ("UNIFORM", "BOOK"):
        return c
    return None


@cashier_bp.route("/cashier/reservations")
def cashier_reservations():
    if not _require_cashier():
        return redirect(url_for("auth.login"))

    branch_id = session.get("branch_id")
    conn = get_db_connection()
    cur = None
    try:
        cur = conn.cursor()

        cur.execute("""
            SELECT
                r.reservation_id,          -- 0
                COALESCE(u.username, '') AS username,   -- 1
                r.student_user_id,         -- 2
                r.student_grade_level,     -- 3
                COALESCE(
                CONCAT_WS(' ',
                    e.student_first_name,
                    e.student_middle_name,
                    e.student_last_name
                    ),
                    svp.student_name,
                    u.username,
                    ''
                ) AS full_name,          -- 4
                COALESCE(r.student_grade_level, svp.grade_level) AS grade_level, -- 5
                NULL AS strand,            -- 6
                r.status,                  -- 7
                r.created_at,              -- 8
                CASE
                  WHEN r.reserved_by_user_id IS NOT NULL
                       AND reserved_by.role = 'parent'
                  THEN 'parent'
                  ELSE 'student'
                END AS reserved_by_role,   -- 9
                CASE
                  WHEN r.reserved_by_user_id IS NOT NULL
                       AND reserved_by.role = 'parent'
                  THEN COALESCE(
                    svp.guardian_name,
                    reserved_by.username
                  )
                  ELSE NULL
                END AS parent_name,        -- 10
                svp.relationship,          -- 11
                (
                    SELECT STRING_AGG(DISTINCT UPPER(ii.category), ',')
                    FROM reservation_items ri
                    JOIN inventory_items ii ON ri.item_id = ii.item_id
                    WHERE ri.reservation_id = r.reservation_id
                ) AS item_categories       -- 12
            FROM reservations r
            LEFT JOIN users u ON u.user_id = r.student_user_id
            LEFT JOIN student_accounts sa ON sa.username = u.username
            LEFT JOIN enrollments e ON e.enrollment_id = sa.enrollment_id
            LEFT JOIN users reserved_by ON reserved_by.user_id = r.reserved_by_user_id
            LEFT JOIN LATERAL (
                SELECT
    CONCAT_WS(' ',
        e2.student_first_name,
        e2.student_middle_name,
        e2.student_last_name
    ) AS student_name,
    e2.grade_level,
    CONCAT_WS(' ',
        e2.guardian_first_name,
        e2.guardian_middle_name,
        e2.guardian_last_name
    ) AS guardian_name,
    ps2.relationship
                FROM parent_student ps2
                JOIN enrollments e2 ON e2.enrollment_id = ps2.student_id
                WHERE ps2.parent_id = r.reserved_by_user_id
                ORDER BY ps2.student_id
                LIMIT 1
            ) svp ON (reserved_by.role = 'parent')
            WHERE r.branch_id = %s
            ORDER BY r.created_at ASC
        """, (branch_id,))
        rows = cur.fetchall() or []
    finally:
        if cur:
            try:
                cur.close()
            except Exception:
                pass
        conn.close()

    return render_template("cashier_reservations.html", rows=rows)


@cashier_bp.route("/cashier/reservations/<int:reservation_id>")
def cashier_reservation_view(reservation_id):
    if not _require_cashier():
        return redirect(url_for("auth.login"))

    branch_id = session.get("branch_id")
    conn = get_db_connection()
    cur = None
    try:
        cur = conn.cursor()

        cur.execute("""
            SELECT
                r.reservation_id,
                COALESCE(u.username, '') AS username,
                r.student_user_id,
                r.student_grade_level,
                COALESCE(
    CONCAT_WS(' ',
        e.student_first_name,
        e.student_middle_name,
        e.student_last_name
    ),
    svp.student_name,
    u.username,
    ''
) AS full_name,
                COALESCE(r.student_grade_level, svp.grade_level) AS grade_level,
                NULL AS strand,
                r.status,
                r.created_at,
                CASE
                  WHEN r.reserved_by_user_id IS NOT NULL
                       AND reserved_by.role = 'parent'
                  THEN 'parent'
                  ELSE 'student'
                END AS reserved_by_role,
                CASE
                  WHEN r.reserved_by_user_id IS NOT NULL
                       AND reserved_by.role = 'parent'
                  THEN COALESCE(
                    svp.guardian_name,
                    reserved_by.username
                  )
                  ELSE NULL
                END AS parent_name,
                svp.relationship
            FROM reservations r
            LEFT JOIN users u ON u.user_id = r.student_user_id
            LEFT JOIN student_accounts sa ON sa.username = u.username
            LEFT JOIN enrollments e ON e.enrollment_id = sa.enrollment_id
            LEFT JOIN users reserved_by ON reserved_by.user_id = r.reserved_by_user_id
            LEFT JOIN LATERAL (
                SELECT
    CONCAT_WS(' ',
        e2.student_first_name,
        e2.student_middle_name,
        e2.student_last_name
    ) AS student_name,
    e2.grade_level,
    CONCAT_WS(' ',
        e2.guardian_first_name,
        e2.guardian_middle_name,
        e2.guardian_last_name
    ) AS guardian_name,
    ps2.relationship
                FROM parent_student ps2
                JOIN enrollments e2 ON e2.enrollment_id = ps2.student_id
                WHERE ps2.parent_id = r.reserved_by_user_id
                ORDER BY ps2.student_id
                LIMIT 1
            ) svp ON (reserved_by.role = 'parent')
            WHERE r.reservation_id = %s AND r.branch_id = %s
            LIMIT 1
        """, (reservation_id, branch_id))
        header = cur.fetchone()
        if not header:
            return render_template("template_missing.html", missing="Reservation not found")

        # 1) Get available categories for this reservation (so UI can show BOOK + UNIFORM)
        cur.execute("""
            SELECT DISTINCT ii.category
            FROM reservation_items ri
            JOIN inventory_items ii ON ii.item_id = ri.item_id
            WHERE ri.reservation_id = %s
            ORDER BY ii.category
        """, (reservation_id,))
        categories = [row[0] for row in (cur.fetchall() or []) if row and row[0]]

        # 2) Determine selected category (default UNIFORM if present)
        selected_category = _normalize_category(request.args.get("category"))
        if not selected_category:
            if "UNIFORM" in categories:
                selected_category = "UNIFORM"
            elif categories:
                # fallback to first available category in DB
                selected_category = str(categories[0]).upper()
            else:
                selected_category = None

        # 3) Compute grand total (ALL categories)
        cur.execute("""
            SELECT COALESCE(SUM(ri.line_total), 0)
            FROM reservation_items ri
            WHERE ri.reservation_id = %s
        """, (reservation_id,))
        grand_total = cur.fetchone()[0] or 0

        # 4) Fetch items (filtered by selected_category, like your UI tabs/filter)
        if selected_category:
            cur.execute("""
                SELECT ii.item_name, ri.qty,
                       COALESCE(NULLIF(TRIM(ri.size_label), ''), ii.publisher, ii.size_label) AS display_label,
                       ri.unit_price, ri.line_total, ii.category, ii.image_url
                FROM reservation_items ri
                JOIN inventory_items ii ON ii.item_id = ri.item_id
                WHERE ri.reservation_id = %s
                  AND UPPER(ii.category) = %s
                ORDER BY ii.item_name
            """, (reservation_id, selected_category))
        else:
            # no items / no category
            cur.execute("""
                SELECT ii.item_name, ri.qty,
                       COALESCE(NULLIF(TRIM(ri.size_label), ''), ii.publisher, ii.size_label) AS display_label,
                       ri.unit_price, ri.line_total, ii.category, ii.image_url
                FROM reservation_items ri
                JOIN inventory_items ii ON ii.item_id = ri.item_id
                WHERE ri.reservation_id = %s
                ORDER BY ii.category, ii.item_name
            """, (reservation_id,))

        items = cur.fetchall() or []
        total = sum(item[4] for item in items)  # filtered total (based on selected_category)

    finally:
        if cur:
            try:
                cur.close()
            except Exception:
                pass
        conn.close()

    return render_template(
        "cashier_reservation_view.html",
        header=header,
        items=items,
        total=total,                 # filtered total
        grand_total=grand_total,     # total for ALL items
        categories=categories,       # ['BOOK','UNIFORM', ...]
        selected_category=selected_category
    )


@cashier_bp.route("/cashier/reservations/<int:reservation_id>/mark-paid", methods=["POST"])
def cashier_mark_paid(reservation_id):
    if not _require_cashier():
        return redirect(url_for("auth.login"))

    branch_id = session.get("branch_id")
    if not is_branch_active(branch_id):
        flash("This branch is currently deactivated. Changes to reservations are not allowed.", "error")
        return redirect(url_for("cashier.cashier_reservation_view", reservation_id=reservation_id))
    conn = get_db_connection()
    cur = None
    try:
        cur = conn.cursor()
        cur.execute("""
            UPDATE reservations
            SET status = 'PAID', paid_at = NOW()
            WHERE reservation_id = %s AND branch_id = %s AND status = 'RESERVED'
        """, (reservation_id, branch_id))
        conn.commit()
    except Exception:
        try:
            conn.rollback()
        except Exception:
            pass
        raise
    finally:
        if cur:
            try:
                cur.close()
            except Exception:
                pass
        conn.close()

    return redirect(url_for("cashier.cashier_reservation_view", reservation_id=reservation_id))


@cashier_bp.route("/cashier/reservations/<int:reservation_id>/mark-claimed", methods=["POST"])
def cashier_mark_claimed(reservation_id):
    if not _require_cashier():
        return redirect(url_for("auth.login"))

    branch_id = session.get("branch_id")
    if not is_branch_active(branch_id):
        flash("This branch is currently deactivated. Changes to reservations are not allowed.", "error")
        return redirect(url_for("cashier.cashier_reservation_view", reservation_id=reservation_id))
    conn = get_db_connection()
    cur = None
    try:
        cur = conn.cursor()

        cur.execute("""
            SELECT status
            FROM reservations
            WHERE reservation_id = %s AND branch_id = %s
            FOR UPDATE
        """, (reservation_id, branch_id))
        r = cur.fetchone()
        if not r:
            raise Exception("Reservation not found.")
        if r[0] not in ("PAID", "RESERVED"):
            raise Exception("Reservation must be RESERVED or PAID.")

        cur.execute("SELECT item_id, qty FROM reservation_items WHERE reservation_id = %s", (reservation_id,))
        lines = cur.fetchall() or []

        for item_id, qty in lines:
            cur.execute("""
                SELECT stock_total, reserved_qty
                FROM inventory_items
                WHERE item_id = %s AND branch_id = %s
                FOR UPDATE
            """, (item_id, branch_id))
            it = cur.fetchone()
            if not it:
                raise Exception("Item not found.")

            stock_total, reserved_qty = int(it[0] or 0), int(it[1] or 0)

            if qty > reserved_qty or qty > stock_total:
                raise Exception("Stock mismatch.")

            # ✅ FIX: include branch_id in WHERE to avoid updating other branches
            cur.execute("""
                UPDATE inventory_items
                SET stock_total = stock_total - %s,
                    reserved_qty = reserved_qty - %s
                WHERE item_id = %s AND branch_id = %s
            """, (qty, qty, item_id, branch_id))

        cur.execute("""
            UPDATE reservations
            SET status = 'CLAIMED', claimed_at = NOW()
            WHERE reservation_id = %s AND branch_id = %s
        """, (reservation_id, branch_id))

        conn.commit()

    except Exception:
        try:
            conn.rollback()
        except Exception:
            pass
        raise
    finally:
        if cur:
            try:
                cur.close()
            except Exception:
                pass
        conn.close()

    return redirect(url_for("cashier.cashier_reservation_view", reservation_id=reservation_id))


@cashier_bp.route("/cashier/reservations/<int:reservation_id>/cancel", methods=["POST"])
def cashier_cancel_reservation(reservation_id):
    if not _require_cashier():
        return redirect(url_for("auth.login"))

    branch_id = session.get("branch_id")
    if not is_branch_active(branch_id):
        flash("This branch is currently deactivated. Changes to reservations are not allowed.", "error")
        return redirect(url_for("cashier.cashier_reservation_view", reservation_id=reservation_id))
    conn = get_db_connection()
    cur = None
    try:
        cur = conn.cursor()

        cur.execute("""
            SELECT status
            FROM reservations
            WHERE reservation_id = %s AND branch_id = %s
            FOR UPDATE
        """, (reservation_id, branch_id))
        r = cur.fetchone()
        if not r:
            raise Exception("Reservation not found.")
        if r[0] != "RESERVED":
            raise Exception("Only RESERVED can be cancelled.")

        cur.execute("SELECT item_id, qty FROM reservation_items WHERE reservation_id = %s", (reservation_id,))
        lines = cur.fetchall() or []

        for item_id, qty in lines:
            cur.execute("""
                UPDATE inventory_items
                SET reserved_qty = GREATEST(reserved_qty - %s, 0)
                WHERE item_id = %s AND branch_id = %s
            """, (qty, item_id, branch_id))

        cur.execute("""
            UPDATE reservations
            SET status = 'CANCELLED', cancelled_at = NOW()
            WHERE reservation_id = %s AND branch_id = %s
        """, (reservation_id, branch_id))

        # --- UPDATE BILLING IF EXISTS ---
        # 1. Calculate removal totals and find enrollment_id
        cur.execute("""
            SELECT
                COALESCE(r.enrollment_id, sa.enrollment_id, u_link.enrollment_id) as enrollment_id,
                COALESCE(SUM(CASE WHEN UPPER(ii.category) = 'BOOK' THEN ri.line_total ELSE 0 END), 0) as book_total,
                COALESCE(SUM(CASE WHEN UPPER(ii.category) = 'UNIFORM' THEN ri.line_total ELSE 0 END), 0) as uniform_total,
                COALESCE(SUM(ri.line_total), 0) as grand_total
            FROM reservations r
            JOIN reservation_items ri ON r.reservation_id = ri.reservation_id
            JOIN inventory_items ii ON ri.item_id = ii.item_id
            -- Try to find the enrollment ID through multiple paths
            LEFT JOIN users u ON r.student_user_id = u.user_id
            LEFT JOIN student_accounts sa ON u.username = sa.username
            LEFT JOIN enrollments u_link ON u_link.user_id = r.student_user_id
            WHERE r.reservation_id = %s
            GROUP BY COALESCE(r.enrollment_id, sa.enrollment_id, u_link.enrollment_id)
        """, (reservation_id,))
        res_data = cur.fetchone()

        if res_data and res_data['enrollment_id']:
            e_id = res_data['enrollment_id']
            b_rem = Decimal(str(res_data['book_total']))
            u_rem = Decimal(str(res_data['uniform_total']))
            g_rem = Decimal(str(res_data['grand_total']))

            # Update fees and total
            cur.execute("""
                UPDATE billing
                SET
                    books_fee = GREATEST(books_fee - %s, 0),
                    uniform_fee = GREATEST(uniform_fee - %s, 0),
                    total_amount = GREATEST(total_amount - %s, 0)
                WHERE enrollment_id = %s AND branch_id = %s
                RETURNING total_amount, amount_paid
            """, (b_rem, u_rem, g_rem, e_id, branch_id))
            bill_update = cur.fetchone()

            if bill_update:
                new_total = bill_update['total_amount']
                paid = bill_update['amount_paid']
                new_balance = max(new_total - paid, 0)
                # Recalculate status
                new_status = 'paid' if new_balance == 0 and new_total > 0 else ('pending' if paid == 0 else 'partial')
                if new_total == 0 and paid == 0:
                    new_status = 'pending'

                cur.execute("""
                    UPDATE billing
                    SET balance = %s, status = %s
                    WHERE enrollment_id = %s AND branch_id = %s
                """, (new_balance, new_status, e_id, branch_id))

        conn.commit()

    except Exception:
        try:
            conn.rollback()
        except Exception:
            pass
        raise
    finally:
        if cur:
            try:
                cur.close()
            except Exception:
                pass
        conn.close()

    return redirect(url_for("cashier.cashier_reservation_view", reservation_id=reservation_id))


@cashier_bp.route("/cashier/reservations/<int:reservation_id>/receipt")
def reservation_receipt(reservation_id):
    if not _require_cashier():
        return redirect(url_for("auth.login"))

    branch_id = session.get("branch_id")
    conn = get_db_connection()
    cur = None
    try:
        cur = conn.cursor()

        cur.execute("""
            SELECT
                r.reservation_id,                       -- 0
                COALESCE(u.username, '') AS username,   -- 1
                r.student_grade_level,                  -- 2
                r.status,                               -- 3
                r.created_at,                           -- 4
                r.claimed_at,                           -- 5
                b.branch_name,                          -- 6
                CASE
                  WHEN r.reserved_by_user_id IS NOT NULL
                       AND reserved_by.role = 'parent'
                  THEN 'parent'
                  ELSE 'student'
                END AS reserved_by_role,                -- 7
                CASE
                  WHEN r.reserved_by_user_id IS NOT NULL
                       AND reserved_by.role = 'parent'
                  THEN COALESCE(svp.guardian_name, reserved_by.username)
                  ELSE NULL
                END AS parent_name,                     -- 8
                svp.relationship,                       -- 9
                COALESCE(
    CONCAT_WS(' ',
        e.student_first_name,
        e.student_middle_name,
        e.student_last_name
    ),
    svp.student_name,
    u.username,
    ''
) AS student_name -- 10
            FROM reservations r
            LEFT JOIN users u ON u.user_id = r.student_user_id
            LEFT JOIN student_accounts sa ON sa.username = u.username
            LEFT JOIN enrollments e ON e.enrollment_id = sa.enrollment_id
            LEFT JOIN branches b ON b.branch_id = r.branch_id
            LEFT JOIN users reserved_by ON reserved_by.user_id = r.reserved_by_user_id
            LEFT JOIN LATERAL (
                SELECT
    CONCAT_WS(' ',
        e2.student_first_name,
        e2.student_middle_name,
        e2.student_last_name
    ) AS student_name,
    e2.grade_level,
    CONCAT_WS(' ',
        e2.guardian_first_name,
        e2.guardian_middle_name,
        e2.guardian_last_name
    ) AS guardian_name,
    ps2.relationship
                FROM parent_student ps2
                JOIN enrollments e2 ON e2.enrollment_id = ps2.student_id
                WHERE ps2.parent_id = r.reserved_by_user_id
                ORDER BY ps2.student_id
                LIMIT 1
            ) svp ON (reserved_by.role = 'parent')
            WHERE r.reservation_id = %s AND r.branch_id = %s
            LIMIT 1
        """, (reservation_id, branch_id))
        header = cur.fetchone()

        if not header:
            return render_template("template_missing.html", missing="Receipt not found"), 404

        # NOTE: status is header[3]
        if header[3] != "CLAIMED":
            return render_template("template_missing.html", missing="Receipt only for claimed"), 403

        cur.execute("""
            SELECT ii.item_name, ri.qty,
                   COALESCE(NULLIF(TRIM(ri.size_label), ''), ii.publisher, ii.size_label) AS display_label,
                   ri.unit_price, ri.line_total, ii.category
            FROM reservation_items ri
            JOIN inventory_items ii ON ii.item_id = ri.item_id
            WHERE ri.reservation_id = %s
            ORDER BY ii.category, ii.item_name
        """, (reservation_id,))
        items = cur.fetchall() or []

        total = sum(item[4] for item in items)

        return render_template("reservation_receipt.html", header=header, items=items, total=total, now=datetime.now)
    finally:
        if cur:
            try:
                cur.close()
            except Exception:
                pass
        conn.close()
        

# =======================
# CSV EXPORT ROUTES
# =======================

@cashier_bp.route("/cashier/reservations/export")
def export_reservations_excel():
    """Export ALL reservations for this branch as a styled Excel file."""
    if not _require_cashier():
        return redirect(url_for("auth.login"))

    branch_id = session.get("branch_id")
    conn = get_db_connection()
    cur = None
    try:
        cur = conn.cursor()
        
        # Get Branch and Admin info for header
        cur.execute("""
            SELECT b.branch_name, u.full_name
            FROM branches b
            LEFT JOIN users u ON u.branch_id = b.branch_id AND u.role = 'branch_admin'
            WHERE b.branch_id = %s
            LIMIT 1
        """, (branch_id,))
        branch_info = cur.fetchone()
        branch_name = branch_info[0] if branch_info else "Unknown Branch"
        admin_name = branch_info[1] if branch_info else "System Administrator"

        cur.execute("""
            SELECT
                r.reservation_id,
                COALESCE(u.username, '') AS username,
                COALESCE(
    CONCAT_WS(' ',
        e.student_first_name,
        e.student_middle_name,
        e.student_last_name
    ),
    svp.student_name,
    u.username,
    ''
) AS full_name,
                COALESCE(r.student_grade_level, svp.grade_level) AS grade_level,
                r.status,
                r.created_at,
                CASE
                  WHEN r.reserved_by_user_id IS NOT NULL AND reserved_by.role = 'parent' THEN 'Parent'
                  ELSE 'Student'
                END AS reserved_by_role,
                CASE
                  WHEN r.reserved_by_user_id IS NOT NULL AND reserved_by.role = 'parent' THEN COALESCE(svp.guardian_name, reserved_by.username)
                  ELSE NULL
                END AS parent_name,
                (
                    SELECT STRING_AGG(DISTINCT UPPER(ii.category), ', ')
                    FROM reservation_items ri
                    JOIN inventory_items ii ON ri.item_id = ii.item_id
                    WHERE ri.reservation_id = r.reservation_id
                ) AS item_categories,
                COALESCE(
                    (SELECT SUM(ri.line_total)
                     FROM reservation_items ri
                     WHERE ri.reservation_id = r.reservation_id), 0
                ) AS grand_total
            FROM reservations r
            LEFT JOIN users u ON u.user_id = r.student_user_id
            LEFT JOIN student_accounts sa ON sa.username = u.username
            LEFT JOIN enrollments e ON e.enrollment_id = sa.enrollment_id
            LEFT JOIN users reserved_by ON reserved_by.user_id = r.reserved_by_user_id
            LEFT JOIN LATERAL (
                SELECT
    CONCAT_WS(' ',
        e2.student_first_name,
        e2.student_middle_name,
        e2.student_last_name
    ) AS student_name,
    e2.grade_level,
    CONCAT_WS(' ',
        e2.guardian_first_name,
        e2.guardian_middle_name,
        e2.guardian_last_name
    ) AS guardian_name,
    ps2.relationship
                FROM parent_student ps2
                JOIN enrollments e2 ON e2.enrollment_id = ps2.student_id
                WHERE ps2.parent_id = r.reserved_by_user_id
                ORDER BY ps2.student_id LIMIT 1
            ) svp ON (reserved_by.role = 'parent')
            WHERE r.branch_id = %s
            ORDER BY r.created_at DESC
        """, (branch_id,))
        rows = cur.fetchall() or []
    finally:
        if cur:
            try: cur.close()
            except Exception: pass
        conn.close()

    # Create Data
    data = []
    for r in rows:
        data.append({
            "Reservation ID": f"RES-{r[0]:04d}",
            "Student Username": r[1] or "",
            "Full Name": r[2] or r[1],
            "Grade Level": r[3] or "",
            "Status": r[4] or "",
            "Date Reserved": r[5].strftime("%Y-%m-%d %H:%M") if r[5] else "",
            "Reserved By": r[6] or "Student",
            "Parent Name": r[7] or "",
            "Item Categories": r[8] or "",
            "Grand Total (PHP)": float(r[9] or 0)
        })

    df = pd.DataFrame(data)
    
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, startrow=4, sheet_name='Reservations')
        workbook = writer.book
        worksheet = writer.sheets['Reservations']

        # Styling
        blue_fill = PatternFill(start_color='1A3A8F', end_color='1A3A8F', fill_type='solid')
        white_font = Font(color='FFFFFF', bold=True, name='Arial', size=10)
        gold_font = Font(color='1A3A8F', bold=True, size=16, name='Arial') # Changed to blue for better print contrast
        header_font = Font(bold=True, size=11, name='Arial')
        thin_border = Border(
            left=Side(style='thin', color='CBD5E1'),
            right=Side(style='thin', color='CBD5E1'),
            top=Side(style='thin', color='CBD5E1'),
            bottom=Side(style='thin', color='CBD5E1')
        )

        # Header Info
        worksheet['A1'] = f"LICEO DE MAJAYJAY - {branch_name.upper()}"
        worksheet['A1'].font = gold_font
        worksheet['A2'] = f"OFFICIAL RESERVATION REGISTRY"
        worksheet['A2'].font = header_font
        worksheet['A3'] = f"Administrator: {admin_name} | Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        worksheet['A3'].font = Font(size=9, color='64748B', name='Arial')
        
        # Format Table Headers
        for cell in worksheet[5]:
            cell.fill = blue_fill
            cell.font = white_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border

        # Auto-adjust columns and add borders
        for row in worksheet.iter_rows(min_row=6, max_row=worksheet.max_row):
            for cell in row:
                cell.border = thin_border
                cell.font = Font(name='Arial', size=9)
                cell.alignment = Alignment(vertical='center')

        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except: pass
            adjusted_width = min(max_length + 4, 40)
            worksheet.column_dimensions[column_letter].width = adjusted_width

        # Page Setup for Printing (Long Bond Paper / Legal)
        worksheet.page_setup.orientation = worksheet.ORIENTATION_LANDSCAPE
        worksheet.page_setup.paperSize = worksheet.PAPERSIZE_LEGAL # 8.5 x 14
        worksheet.page_setup.fitToPage = True
        worksheet.page_setup.fitToWidth = 1
        worksheet.page_setup.fitToHeight = 0
        worksheet.print_options.horizontalCentered = True

    today = date.today().strftime("%Y-%m-%d")
    filename = f"reservations_{today}.xlsx"
    output.seek(0)
    return Response(
        output.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@cashier_bp.route("/cashier/reservations/<int:reservation_id>/export")
def export_reservation_detail_excel(reservation_id):
    """Export a single reservation as a styled Excel file."""
    if not _require_cashier():
        return redirect(url_for("auth.login"))

    branch_id = session.get("branch_id")
    conn = get_db_connection()
    cur = None
    try:
        cur = conn.cursor()
        
        # Get Branch and Admin info
        cur.execute("""
            SELECT b.branch_name, u.full_name
            FROM branches b
            LEFT JOIN users u ON u.branch_id = b.branch_id AND u.role = 'branch_admin'
            WHERE b.branch_id = %s
            LIMIT 1
        """, (branch_id,))
        branch_info = cur.fetchone()
        branch_name = branch_info[0] if branch_info else "Unknown Branch"
        admin_name = branch_info[1] if branch_info else "System Administrator"

        cur.execute("""
            SELECT
                r.reservation_id,
                COALESCE(u.username, '') AS username,
                COALESCE(
    CONCAT_WS(' ',
        e.student_first_name,
        e.student_middle_name,
        e.student_last_name
    ),
    svp.student_name,
    u.username,
    ''
) AS full_name,
                COALESCE(r.student_grade_level, svp.grade_level) AS grade_level,
                r.status,
                r.created_at,
                CASE
                  WHEN r.reserved_by_user_id IS NOT NULL AND reserved_by.role = 'parent' THEN 'Parent'
                  ELSE 'Student'
                END AS reserved_by_role,
                CASE
                  WHEN r.reserved_by_user_id IS NOT NULL AND reserved_by.role = 'parent' THEN COALESCE(svp.guardian_name, reserved_by.username)
                  ELSE NULL
                END AS parent_name,
                svp.relationship
            FROM reservations r
            LEFT JOIN users u ON u.user_id = r.student_user_id
            LEFT JOIN student_accounts sa ON sa.username = u.username
            LEFT JOIN enrollments e ON e.enrollment_id = sa.enrollment_id
            LEFT JOIN users reserved_by ON reserved_by.user_id = r.reserved_by_user_id
            LEFT JOIN LATERAL (
                SELECT
    CONCAT_WS(' ',
        e2.student_first_name,
        e2.student_middle_name,
        e2.student_last_name
    ) AS student_name,
    e2.grade_level,
    CONCAT_WS(' ',
        e2.guardian_first_name,
        e2.guardian_middle_name,
        e2.guardian_last_name
    ) AS guardian_name,
    ps2.relationship
                FROM parent_student ps2
                JOIN enrollments e2 ON e2.enrollment_id = ps2.student_id
                WHERE ps2.parent_id = r.reserved_by_user_id
                ORDER BY ps2.student_id LIMIT 1
            ) svp ON (reserved_by.role = 'parent')
            WHERE r.reservation_id = %s AND r.branch_id = %s
            LIMIT 1
        """, (reservation_id, branch_id))
        header = cur.fetchone()
        if not header:
            flash("Reservation not found.", "error")
            return redirect(url_for("cashier.cashier_reservations"))

        cur.execute("""
            SELECT ii.category, ii.item_name, ri.qty,
                   COALESCE(NULLIF(TRIM(ri.size_label), ''), ii.publisher, ii.size_label) AS display_label,
                   ri.unit_price, ri.line_total
            FROM reservation_items ri
            JOIN inventory_items ii ON ii.item_id = ri.item_id
            WHERE ri.reservation_id = %s
            ORDER BY ii.category, ii.item_name
        """, (reservation_id,))
        items = cur.fetchall() or []
        grand_total = sum(float(it[5] or 0) for it in items)
    finally:
        if cur:
            try: cur.close()
            except Exception: pass
        conn.close()

    # Create Item List Data
    item_data = []
    for it in items:
        item_data.append({
            "Category": it[0] or "",
            "Item Name": it[1] or "",
            "Qty": it[2] or 0,
            "Size / Publisher": it[3] or "",
            "Unit Price (PHP)": float(it[4] or 0),
            "Line Total (PHP)": float(it[5] or 0)
        })

    df_items = pd.DataFrame(item_data)
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df_items.to_excel(writer, index=False, startrow=10, sheet_name='Details')
        workbook = writer.book
        worksheet = writer.sheets['Details']

        # Styling
        blue_fill = PatternFill(start_color='1A3A8F', end_color='1A3A8F', fill_type='solid')
        white_font = Font(color='FFFFFF', bold=True, name='Arial', size=10)
        gold_font = Font(color='1A3A8F', bold=True, size=18, name='Arial')
        label_font = Font(bold=True, name='Arial', size=10)
        thin_border = Border(
            left=Side(style='thin', color='CBD5E1'),
            right=Side(style='thin', color='CBD5E1'),
            top=Side(style='thin', color='CBD5E1'),
            bottom=Side(style='thin', color='CBD5E1')
        )

        # Header Section
        worksheet['A1'] = f"LICEO DE MAJAYJAY - {branch_name.upper()}"
        worksheet['A1'].font = gold_font
        worksheet['A2'] = f"INDIVIDUAL RESERVATION REPORT"
        worksheet['A2'].font = Font(bold=True, size=12, name='Arial')
        worksheet['A3'] = f"Branch Administrator: {admin_name} | Exported: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        worksheet['A3'].font = Font(size=9, color='64748B', name='Arial')

        # Summary Section
        worksheet['A5'] = "RESERVATION SUMMARY"
        worksheet['A5'].font = Font(bold=True, size=12, color='1A3A8F', name='Arial')
        
        summary_rows = [
            ("Reservation ID", f"RES-{header[0]:04d}"),
            ("Student Name",   header[2] or header[1] or ""),
            ("Grade Level",    header[3] or ""),
            ("Status",         header[4] or ""),
            ("Date Reserved",  header[5].strftime("%Y-%m-%d %H:%M") if header[5] else "")
        ]
        for i, (label, val) in enumerate(summary_rows):
            worksheet[f'A{6+i}'] = label
            worksheet[f'A{6+i}'].font = label_font
            worksheet[f'B{6+i}'] = val
            worksheet[f'B{6+i}'].font = Font(name='Arial', size=10)

        # Table Header Styling
        for cell in worksheet[11]:
            cell.fill = blue_fill
            cell.font = white_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border

        # Data Row Styling
        for row in worksheet.iter_rows(min_row=12, max_row=11+len(items)):
            for cell in row:
                cell.border = thin_border
                cell.font = Font(name='Arial', size=9)

        # Total Row
        last_row = 12 + len(items)
        worksheet[f'E{last_row}'] = "GRAND TOTAL"
        worksheet[f'E{last_row}'].font = label_font
        worksheet[f'F{last_row}'] = grand_total
        worksheet[f'F{last_row}'].font = Font(bold=True, size=11, name='Arial')
        worksheet[f'F{last_row}'].border = thin_border

        # Auto-adjust columns
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except: pass
            worksheet.column_dimensions[column_letter].width = min(max_length + 4, 50)

        # Page Setup
        worksheet.page_setup.orientation = worksheet.ORIENTATION_PORTRAIT
        worksheet.page_setup.paperSize = worksheet.PAPERSIZE_LEGAL # Long Bond Paper
        worksheet.page_setup.fitToPage = True
        worksheet.page_setup.fitToWidth = 1
        worksheet.page_setup.fitToHeight = 0
        worksheet.print_options.horizontalCentered = True

    filename = f"reservation_RES-{header[0]:04d}.xlsx"
    output.seek(0)
    return Response(
        output.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )



@cashier_bp.route("/cashier/unpaid-report")
def unpaid_report():
    if not _require_cashier():
        return redirect("/")

    db = get_db_connection()
    cursor = db.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

    try:
        branch_id = session.get("branch_id")
        active_year_id = _get_active_year_id(cursor, branch_id)

        if not active_year_id:
            flash("No active school year found. Please contact admin.", "warning")
            return redirect(url_for("cashier.dashboard"))

        # Fetch year label
        cursor.execute("SELECT label FROM school_years WHERE year_id = %s", (active_year_id,))
        year_label = cursor.fetchone()["label"]

        # Fetch students with balances > 0 for this year
        cursor.execute("""
            SELECT
    CONCAT_WS(' ',
        e.student_first_name,
        e.student_middle_name,
        e.student_last_name
    ) AS student_name,
    e.grade_level,
    e.branch_enrollment_no,
    b.bill_id,
    b.total_amount,
    b.amount_paid,
    b.balance,
    b.status
            FROM billing b
            JOIN enrollments e ON b.enrollment_id = e.enrollment_id
            WHERE b.branch_id = %s AND b.year_id = %s AND b.balance > 0
            ORDER BY b.balance DESC
        """, (branch_id, active_year_id))
        unpaid_students = cursor.fetchall()

        # Calculate total outstanding
        total_outstanding = sum(row["balance"] for row in unpaid_students)

        return render_template(
            "cashier_unpaid_report.html",
            unpaid_students=unpaid_students,
            year_label=year_label,
            total_outstanding=total_outstanding
        )
    finally:
        cursor.close()
        db.close()




@cashier_bp.route("/cashier/uniform-orders")
def uniform_orders():
    if not _require_cashier():
        return redirect(url_for("auth.login"))

    branch_id = session.get("branch_id")
    db = get_db_connection()
    cursor = db.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    try:
        active_year_id = _get_active_year_id(cursor, branch_id)
        status_filter = request.args.get("status", "all")
        search = request.args.get("search", "").strip()

        # Grade+section options from enrollments (same as Billing Registry — e.g. Grade 11-GAS)
        cursor.execute("""
            SELECT DISTINCT grade_level
            FROM enrollments
            WHERE branch_id = %s
              AND year_id = %s
              AND grade_level IS NOT NULL
              AND TRIM(grade_level) <> ''
              AND status IN ('approved', 'enrolled')
            ORDER BY grade_level
        """, (branch_id, active_year_id))
        grade_levels = [r["grade_level"] for r in cursor.fetchall()]

        grade_filter_raw = request.args.get("grade")
        if grade_filter_raw is None:
            if "Nursery" in grade_levels:
                grade_filter = "Nursery"
            elif any("nursery" in g.lower() for g in grade_levels):
                grade_filter = next(g for g in grade_levels if "nursery" in g.lower())
            elif grade_levels:
                grade_filter = grade_levels[0]
            else:
                grade_filter = "all"
        else:
            grade_filter = grade_filter_raw.strip()
            if not grade_filter:
                grade_filter = "all"
            elif grade_filter.lower() != "all" and grade_filter not in grade_levels:
                grade_filter = "all"

        params = [branch_id]
        where_extra = ""
        if status_filter != "all":
            where_extra += " AND uo.order_status = %s"
            params.append(status_filter)
        if grade_filter and grade_filter.lower() != "all":
            # Exact match so Grade 11-GAS is not mixed with other Grade 11 sections
            where_extra += " AND e.grade_level = %s"
            params.append(grade_filter)
        if search:
            where_extra += """ AND (
                LOWER(CONCAT_WS(' ', e.student_first_name, e.student_middle_name, e.student_last_name)) LIKE %s
                OR LOWER(uo.order_number) LIKE %s
            )"""
            params += [f"%{search.lower()}%", f"%{search.lower()}%"]

        cursor.execute(f"""
            SELECT
                uo.order_id,
                uo.order_number,
                uo.total_amount,
                uo.payment_status,
                uo.order_status,
                uo.created_at,
                uo.onsite_arrived_at,
                uo.claimed_at,
                uo.enrollment_id,
                uo.bill_id,
                uo.created_by_user_id,
                CONCAT_WS(' ', e.student_first_name, e.student_middle_name, e.student_last_name) AS student_name,
                e.grade_level,
                e.branch_enrollment_no,
                u_creator.role AS creator_role,
                COALESCE(NULLIF(u_creator.full_name, ''), CONCAT_WS(' ', u_creator.first_name, u_creator.last_name), u_creator.username) AS creator_name
            FROM uniform_orders uo
            JOIN enrollments e ON e.enrollment_id = uo.enrollment_id
            LEFT JOIN users u_creator ON u_creator.user_id = uo.created_by_user_id
            WHERE uo.branch_id = %s
            {where_extra}
            ORDER BY uo.created_at DESC
        """, params)
        orders = cursor.fetchall()

        # Stats (branch-wide, not grade-filtered — overview cards)
        cursor.execute("""
            SELECT
                COUNT(*) FILTER (WHERE order_status = 'For Ordering') AS for_ordering,
                COUNT(*) FILTER (WHERE order_status = 'Ready for Claim') AS ready,
                COUNT(*) FILTER (WHERE order_status = 'Claimed') AS claimed,
                COUNT(*) AS total
            FROM uniform_orders WHERE branch_id = %s
        """, (branch_id,))
        stats = cursor.fetchone()

        # Auto-sync parent set prices to sum of child pieces if pieces exist
        cursor.execute("""
            UPDATE inventory_items parent
            SET price = COALESCE((
                SELECT SUM(child.price)
                FROM inventory_items child
                WHERE child.parent_item_id = parent.item_id
                  AND child.branch_id = parent.branch_id
                  AND UPPER(child.category) = 'UNIFORM'
            ), parent.price)
            WHERE parent.branch_id = %s
              AND UPPER(parent.category) = 'UNIFORM'
              AND parent.parent_item_id IS NULL
              AND EXISTS (
                  SELECT 1 FROM inventory_items c
                  WHERE c.parent_item_id = parent.item_id
                    AND c.branch_id = parent.branch_id
                    AND UPPER(c.category) = 'UNIFORM'
              )
        """, (branch_id,))
        db.commit()

        # Fetch uniform SETS (top-level items) for catalog tab
        cursor.execute("""
            SELECT item_id, category, item_name, grade_level, price, image_url, is_active, size_label,
                   COALESCE(size_price_step, 20) AS size_price_step
            FROM inventory_items
            WHERE branch_id = %s AND UPPER(category) = 'UNIFORM'
              AND (parent_item_id IS NULL)
            ORDER BY item_name ASC
        """, (branch_id,))
        catalog_sets = cursor.fetchall()

        # Fetch all pieces (items with a parent_item_id)
        cursor.execute("""
            SELECT item_id, item_name, grade_level, price, size_label, parent_item_id,
                   COALESCE(size_price_step, 20) AS size_price_step
            FROM inventory_items
            WHERE branch_id = %s AND UPPER(category) = 'UNIFORM'
              AND parent_item_id IS NOT NULL
            ORDER BY parent_item_id, item_name ASC
        """, (branch_id,))
        all_pieces = cursor.fetchall()

        # Attach size→price map for display
        for row in catalog_sets:
            row["size_prices"] = size_price_map(row["price"], row["size_label"], row["size_price_step"])
        for row in all_pieces:
            row["size_prices"] = size_price_map(row["price"], row["size_label"], 0)

        # Group pieces by parent_item_id
        pieces_by_set = {}
        for p in all_pieces:
            pid = p['parent_item_id']
            if pid not in pieces_by_set:
                pieces_by_set[pid] = []
            pieces_by_set[pid].append(p)

        return render_template(
            "cashier_uniform_orders.html",
            orders=orders,
            stats=stats,
            status_filter=status_filter,
            grade_filter=grade_filter,
            grade_levels=grade_levels,
            search=search,
            catalog_sets=catalog_sets,
            pieces_by_set=pieces_by_set
        )
    finally:
        cursor.close()
        db.close()


@cashier_bp.route("/cashier/api/sidebar-badges")
def cashier_sidebar_badges():
    """Live counts for cashier sidebar badges (AJAX poll)."""
    if not _require_cashier():
        return jsonify({"error": "Unauthorized"}), 403
    branch_id = session.get("branch_id")
    db = get_db_connection()
    cursor = db.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    try:
        cursor.execute("""
            SELECT COUNT(*) AS n FROM uniform_orders
            WHERE branch_id = %s AND order_status = 'For Ordering'
        """, (branch_id,))
        for_ordering = int(cursor.fetchone()["n"] or 0)

        cursor.execute("""
            SELECT COUNT(*) AS n FROM uniform_orders
            WHERE branch_id = %s AND order_status = 'Ready for Claim'
        """, (branch_id,))
        ready_claim = int(cursor.fetchone()["n"] or 0)

        cursor.execute("""
            SELECT COUNT(*) AS n FROM billing b
            JOIN enrollments e ON e.enrollment_id = b.enrollment_id
            WHERE e.branch_id = %s AND COALESCE(b.balance, 0) > 0
              AND LOWER(COALESCE(b.status, '')) NOT IN ('paid', 'full')
        """, (branch_id,))
        billing_due = int(cursor.fetchone()["n"] or 0)

        cursor.execute("""
            SELECT COUNT(*) AS n FROM reservations r
            WHERE r.branch_id = %s AND UPPER(r.status) = 'RESERVED'
        """, (branch_id,))
        reserved = int(cursor.fetchone()["n"] or 0)

        return jsonify({
            "uniform_for_ordering": for_ordering,
            "uniform_ready": ready_claim,
            "uniform_orders": for_ordering + ready_claim,
            "billing_due": billing_due,
            "reservations": reserved,
        })
    finally:
        cursor.close()
        db.close()


@cashier_bp.route("/cashier/uniform-orders/<int:order_id>/mark-onsite", methods=["POST"])
def uniform_mark_onsite(order_id):
    if not _require_cashier():
        return jsonify({"error": "Unauthorized"}), 403

    branch_id = session.get("branch_id")
    db = get_db_connection()
    cursor = db.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    try:
        cursor.execute("""
            SELECT uo.*, e.student_first_name, e.student_middle_name, e.student_last_name,
                   e.year_id AS enroll_year_id, e.enrollment_id
            FROM uniform_orders uo
            JOIN enrollments e ON e.enrollment_id = uo.enrollment_id
            WHERE uo.order_id = %s AND uo.branch_id = %s
        """, (order_id, branch_id))
        order = cursor.fetchone()
        if not order:
            return jsonify({"error": "Order not found"}), 404
        if order["order_status"] != "For Ordering":
            return jsonify({"error": "Order is already marked onsite or processed"}), 400

        year_id = order["year_id"] or order["enroll_year_id"] or _get_active_year_id(cursor, branch_id)

        # Activate billing entry if not yet linked
        bill_id = order["bill_id"]
        if not bill_id:
            # Check if student already has a billing record
            cursor.execute("""
                SELECT bill_id, uniform_fee FROM billing WHERE enrollment_id = %s
            """, (order["enrollment_id"],))
            existing_bill = cursor.fetchone()
            if existing_bill:
                # Add uniform fee onto existing bill
                new_uniform_fee = float(existing_bill["uniform_fee"] or 0) + float(order["total_amount"])
                cursor.execute("""
                    UPDATE billing
                    SET uniform_fee = %s,
                        total_amount = total_amount + %s,
                        balance = balance + %s,
                        updated_at = NOW()
                    WHERE bill_id = %s
                    RETURNING bill_id
                """, (new_uniform_fee, order["total_amount"], order["total_amount"], existing_bill["bill_id"]))
                bill_id = cursor.fetchone()["bill_id"]
            else:
                cursor.execute("""
                    INSERT INTO billing
                      (enrollment_id, branch_id, year_id, tuition_fee, books_fee, uniform_fee,
                       other_fees, total_amount, amount_paid, balance, status, created_by)
                    VALUES (%s, %s, %s, 0, 0, %s, 0, %s, 0, %s, 'pending', %s)
                    RETURNING bill_id
                """, (
                    order["enrollment_id"], branch_id, year_id,
                    order["total_amount"], order["total_amount"], order["total_amount"],
                    session.get("user_id")
                ))
                bill_id = cursor.fetchone()["bill_id"]

        now = _get_manila_now().replace(tzinfo=None)
        # Ensure updated_at column exists in uniform_orders
        try:
            cursor.execute("ALTER TABLE uniform_orders ADD COLUMN IF NOT EXISTS updated_at TIMESTAMP DEFAULT NOW()")
            cursor.execute("ALTER TABLE uniform_orders ADD COLUMN IF NOT EXISTS claimed_by_user_id INTEGER")
        except Exception:
            pass

        try:
            cursor.execute("""
                UPDATE uniform_orders
                SET order_status = 'Ready for Claim',
                    bill_id = %s,
                    onsite_arrived_at = %s,
                    updated_at = %s
                WHERE order_id = %s
            """, (bill_id, now, now, order_id))
        except Exception:
            cursor.execute("""
                UPDATE uniform_orders
                SET order_status = 'Ready for Claim',
                    bill_id = %s,
                    onsite_arrived_at = %s
                WHERE order_id = %s
            """, (bill_id, now, order_id))

        # Send notification to student
        student_user_id = order.get("student_user_id")
        if not student_user_id:
            cursor.execute("""
                SELECT u.user_id
                FROM student_accounts sa
                JOIN users u ON u.username = sa.username
                WHERE sa.enrollment_id = %s
                LIMIT 1
            """, (order["enrollment_id"],))
            s_row = cursor.fetchone()
            if s_row:
                student_user_id = s_row["user_id"]

        if student_user_id:
            cursor.execute("""
                INSERT INTO student_notifications (student_id, title, message, link)
                VALUES (%s, %s, %s, %s)
            """, (
                student_user_id,
                "Uniform Ready for Claim",
                f"Your uniform order {order['order_number']} is now onsite! Your bill has been activated. Please proceed to the cashier to pay and claim.",
                "/student/reservations"
            ))

        db.commit()

        return jsonify({"success": True, "message": "Marked as Ready for Claim. Bill has been activated."})
    except Exception as e:
        db.rollback()
        return jsonify({"error": str(e)}), 500
    finally:
        cursor.close()
        db.close()


@cashier_bp.route("/cashier/uniform-orders/<int:order_id>/process-claim", methods=["POST"])
def uniform_process_claim(order_id):
    if not _require_cashier():
        return jsonify({"error": "Unauthorized"}), 403

    branch_id = session.get("branch_id")
    db = get_db_connection()
    cursor = db.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    try:
        cursor.execute("""
            SELECT uo.*, b.status AS bill_status, b.balance
            FROM uniform_orders uo
            LEFT JOIN billing b ON b.bill_id = uo.bill_id
            WHERE uo.order_id = %s AND uo.branch_id = %s
        """, (order_id, branch_id))
        order = cursor.fetchone()
        if not order:
            return jsonify({"error": "Order not found"}), 404
        if order["order_status"] not in ("Ready for Claim",):
            return jsonify({"error": "Order is not yet Ready for Claim"}), 400
        if (order.get("payment_status") or "").lower() != "paid":
            return jsonify({"error": "Uniform order is Unpaid. Payment must be collected at Cashier before release."}), 400

        now = _get_manila_now().replace(tzinfo=None)
        try:
            cursor.execute("""
                UPDATE uniform_orders
                SET order_status = 'Claimed',
                    payment_status = 'Paid',
                    claimed_at = %s,
                    claimed_by_user_id = %s,
                    updated_at = %s
                WHERE order_id = %s
            """, (now, session.get("user_id"), now, order_id))
        except Exception:
            cursor.execute("""
                UPDATE uniform_orders
                SET order_status = 'Claimed',
                    payment_status = 'Paid',
                    claimed_at = %s
                WHERE order_id = %s
            """, (now, order_id))

        student_user_id = order.get("student_user_id")
        if not student_user_id:
            cursor.execute("""
                SELECT u.user_id
                FROM student_accounts sa
                JOIN users u ON u.username = sa.username
                WHERE sa.enrollment_id = %s
                LIMIT 1
            """, (order["enrollment_id"],))
            s_row = cursor.fetchone()
            if s_row:
                student_user_id = s_row["user_id"]
        if student_user_id:
            cursor.execute("""
                INSERT INTO student_notifications (student_id, title, message, link)
                VALUES (%s, %s, %s, %s)
            """, (
                student_user_id,
                "Uniform Claimed",
                f"Your uniform order {order['order_number']} has been claimed. Thank you!",
                "/student/reservations"
            ))

        db.commit()
        return jsonify({"success": True, "message": "Order marked as Claimed."})
    except Exception as e:
        db.rollback()
        return jsonify({"error": str(e)}), 500
    finally:
        cursor.close()
        db.close()


@cashier_bp.route("/cashier/uniform-orders/<int:order_id>/items")
def uniform_order_items(order_id):
    if not _require_cashier():
        return jsonify({"error": "Unauthorized"}), 403

    branch_id = session.get("branch_id")
    db = get_db_connection()
    cursor = db.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    try:
        cursor.execute("""
            SELECT uo.created_by_user_id, u_creator.role AS creator_role,
                   COALESCE(NULLIF(u_creator.full_name, ''), CONCAT_WS(' ', u_creator.first_name, u_creator.last_name), u_creator.username) AS creator_name
            FROM uniform_orders uo
            LEFT JOIN users u_creator ON u_creator.user_id = uo.created_by_user_id
            WHERE uo.order_id = %s AND uo.branch_id = %s
        """, (order_id, branch_id))
        meta = cursor.fetchone() or {}

        cursor.execute("""
            SELECT uoi.item_name, uoi.size_label, uoi.quantity, uoi.unit_price, uoi.line_total
            FROM uniform_order_items uoi
            JOIN uniform_orders uo ON uo.order_id = uoi.order_id
            WHERE uoi.order_id = %s AND uo.branch_id = %s
            ORDER BY uoi.item_id
        """, (order_id, branch_id))
        items = cursor.fetchall()
        return jsonify({
            "creator_role": meta.get("creator_role"),
            "creator_name": meta.get("creator_name"),
            "items": [dict(i) for i in items]
        })
    finally:
        cursor.close()
        db.close()


def _sync_parent_set_price(cursor, parent_item_id, branch_id):
    """Auto-calculate parent uniform set base price as the sum of its child pieces."""
    if not parent_item_id:
        return
    cursor.execute("""
        SELECT COALESCE(SUM(price), 0) AS total_pieces_price, COUNT(*) AS piece_count
        FROM inventory_items
        WHERE parent_item_id = %s AND branch_id = %s AND UPPER(category) = 'UNIFORM'
    """, (int(parent_item_id), branch_id))
    r = cursor.fetchone()
    if r and r["piece_count"] > 0:
        total_price = float(r["total_pieces_price"] or 0)
        cursor.execute("""
            UPDATE inventory_items
            SET price = %s
            WHERE item_id = %s AND branch_id = %s AND UPPER(category) = 'UNIFORM'
        """, (total_price, int(parent_item_id), branch_id))


@cashier_bp.route("/cashier/uniform-catalog/update-price", methods=["POST"])
def uniform_catalog_update_price():
    if not _require_cashier():
        return jsonify({"error": "Unauthorized"}), 403

    branch_id = session.get("branch_id")
    item_id = request.form.get("item_id")
    new_price_str = (request.form.get("new_price") or "").strip()
    size_price_step_str = (request.form.get("size_price_step") or str(DEFAULT_SIZE_PRICE_STEP)).strip()

    if not item_id or not new_price_str:
        return jsonify({"error": "Missing item_id or price"}), 400

    try:
        new_price = float(new_price_str)
        size_price_step = float(size_price_step_str or DEFAULT_SIZE_PRICE_STEP)
    except ValueError:
        return jsonify({"error": "Price must be a valid number"}), 400

    if new_price < 0:
        return jsonify({"error": "Price cannot be negative"}), 400
    if size_price_step < 0 or size_price_step > 999.99:
        return jsonify({"error": "Price increment per size must be between ₱0.00 and ₱999.99 (max 3 digits)"}), 400

    db = get_db_connection()
    cursor = db.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    try:
        cursor.execute(
            "SELECT parent_item_id, item_name FROM inventory_items WHERE item_id = %s AND branch_id = %s",
            (item_id, branch_id)
        )
        row = cursor.fetchone()
        if not row:
            return jsonify({"error": "Item not found"}), 404

        is_piece = row["parent_item_id"] is not None
        if is_piece:
            if new_price > 999.99:
                return jsonify({"error": "Individual piece price cannot exceed ₱999.99 (maximum 3 digits)"}), 400
        else:
            if new_price > 9999.99:
                return jsonify({"error": "Full set price cannot exceed ₱9,999.99 (maximum 4 digits)"}), 400

        cursor.execute("""
            UPDATE inventory_items
            SET price = %s, size_price_step = %s
            WHERE item_id = %s AND branch_id = %s AND UPPER(category) = 'UNIFORM'
        """, (new_price, size_price_step, item_id, branch_id))

        if is_piece and row["parent_item_id"]:
            _sync_parent_set_price(cursor, row["parent_item_id"], branch_id)

        db.commit()
        return jsonify({"success": True, "message": "Price updated successfully"})
    except Exception as e:
        db.rollback()
        return jsonify({"error": str(e)}), 500
    finally:
        cursor.close()
        db.close()


@cashier_bp.route("/cashier/uniform-catalog/add", methods=["POST"])
def uniform_catalog_add():
    """Add a new Uniform SET (top-level catalog item, no parent)."""
    if not _require_cashier():
        return jsonify({"error": "Unauthorized"}), 403

    branch_id = session.get("branch_id")
    item_name = request.form.get("item_name", "").strip()
    grade_level = request.form.get("grade_level", "All Grades").strip()
    price = request.form.get("price", "0").strip()
    size_label = request.form.get("size_label", "XS, S, M, L, XL, XXL, XXXL").strip()
    size_price_step = request.form.get("size_price_step", str(DEFAULT_SIZE_PRICE_STEP)).strip()
    image_url = request.form.get("image_url", "").strip()

    if not item_name:
        return jsonify({"error": "Item name is required"}), 400

    try:
        price_val = float(price or 0)
        step_val = float(size_price_step or DEFAULT_SIZE_PRICE_STEP)
    except ValueError:
        return jsonify({"error": "Price must be a valid number"}), 400

    if price_val < 0:
        return jsonify({"error": "Price cannot be negative"}), 400
    if price_val > 9999.99:
        return jsonify({"error": "Full set price cannot exceed ₱9,999.99 (maximum 4 digits)"}), 400
    if step_val < 0 or step_val > 999.99:
        return jsonify({"error": "Price increment per size must be between ₱0.00 and ₱999.99 (max 3 digits)"}), 400

    # Normalize size list order
    size_label = ", ".join(parse_size_list(size_label))

    db = get_db_connection()
    cursor = db.cursor()
    try:
        cursor.execute("""
            INSERT INTO inventory_items
              (branch_id, category, item_name, grade_level, price, size_label, image_url,
               is_active, parent_item_id, is_set_piece, size_price_step)
            VALUES (%s, 'UNIFORM', %s, %s, %s, %s, %s, TRUE, NULL, FALSE, %s)
        """, (
            branch_id, item_name, grade_level, price_val, size_label,
            image_url or None, step_val
        ))
        db.commit()
        return jsonify({"success": True, "message": "Uniform set added successfully"})
    except Exception as e:
        db.rollback()
        return jsonify({"error": str(e)}), 500
    finally:
        cursor.close()
        db.close()


@cashier_bp.route("/cashier/uniform-catalog/add-piece", methods=["POST"])
def uniform_catalog_add_piece():
    """Add an individual PIECE under a specific uniform set."""
    if not _require_cashier():
        return jsonify({"error": "Unauthorized"}), 403

    branch_id = session.get("branch_id")
    item_name = request.form.get("item_name", "").strip()
    parent_item_id = request.form.get("parent_item_id", "").strip()
    price = request.form.get("price", "0").strip()
    size_label = request.form.get("size_label", "XS, S, M, L, XL, XXL, XXXL").strip()
    size_price_step = request.form.get("size_price_step", str(DEFAULT_SIZE_PRICE_STEP)).strip()

    if not item_name or not parent_item_id:
        return jsonify({"error": "Item name and parent set are required"}), 400

    try:
        price_val = float(price or 0)
        step_val = float(size_price_step or DEFAULT_SIZE_PRICE_STEP)
    except ValueError:
        return jsonify({"error": "Price must be a valid number"}), 400

    if price_val < 0:
        return jsonify({"error": "Price cannot be negative"}), 400
    if price_val > 999.99:
        return jsonify({"error": "Individual piece price cannot exceed ₱999.99 (maximum 3 digits)"}), 400
    if step_val < 0 or step_val > 999.99:
        return jsonify({"error": "Price increment per size must be between ₱0.00 and ₱999.99 (max 3 digits)"}), 400

    size_label = ", ".join(parse_size_list(size_label))

    db = get_db_connection()
    cursor = db.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    try:
        # Inherit grade_level from parent set
        cursor.execute(
            "SELECT grade_level FROM inventory_items WHERE item_id = %s AND branch_id = %s",
            (int(parent_item_id), branch_id)
        )
        parent = cursor.fetchone()
        if not parent:
            return jsonify({"error": "Parent set not found"}), 404

        # Enforce max 2 pieces per set (e.g. Polo & Pants)
        cursor.execute(
            "SELECT COUNT(*) AS piece_count FROM inventory_items WHERE parent_item_id = %s AND branch_id = %s",
            (int(parent_item_id), branch_id)
        )
        p_row = cursor.fetchone()
        if p_row and int(p_row["piece_count"] or 0) >= 2:
            return jsonify({"error": "Maximum 2 pieces allowed per uniform set (e.g. Polo and Pants)."}), 400

        cursor.execute("""
            INSERT INTO inventory_items
              (branch_id, category, item_name, grade_level, price, size_label,
               is_active, parent_item_id, is_set_piece, size_price_step)
            VALUES (%s, 'UNIFORM', %s, %s, %s, %s, TRUE, %s, TRUE, %s)
        """, (
            branch_id, item_name, parent['grade_level'], price_val, size_label,
            int(parent_item_id), step_val
        ))

        _sync_parent_set_price(cursor, int(parent_item_id), branch_id)

        db.commit()
        return jsonify({"success": True, "message": "Piece added to set successfully"})
    except Exception as e:
        db.rollback()
        return jsonify({"error": str(e)}), 500
    finally:
        cursor.close()
        db.close()


@cashier_bp.route("/cashier/uniform-catalog/delete-piece", methods=["POST"])
def uniform_catalog_delete_piece():
    """Delete an individual uniform PIECE (not a full set)."""
    if not _require_cashier():
        return jsonify({"error": "Unauthorized"}), 403

    branch_id = session.get("branch_id")
    item_id = request.form.get("item_id", "").strip()
    if not item_id:
        return jsonify({"error": "Missing item_id"}), 400

    db = get_db_connection()
    cursor = db.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    try:
        cursor.execute("""
            SELECT item_id, item_name, parent_item_id, COALESCE(is_set_piece, FALSE) AS is_set_piece
            FROM inventory_items
            WHERE item_id = %s AND branch_id = %s AND UPPER(category) = 'UNIFORM'
        """, (int(item_id), branch_id))
        row = cursor.fetchone()
        if not row:
            return jsonify({"error": "Piece not found"}), 404
        if not row["parent_item_id"] and not row["is_set_piece"]:
            return jsonify({"error": "Only individual pieces can be deleted here"}), 400

        parent_id = row["parent_item_id"]

        cursor.execute("""
            DELETE FROM inventory_items
            WHERE item_id = %s AND branch_id = %s
              AND (parent_item_id IS NOT NULL OR COALESCE(is_set_piece, FALSE) = TRUE)
        """, (int(item_id), branch_id))

        if parent_id:
            _sync_parent_set_price(cursor, parent_id, branch_id)

        db.commit()
        return jsonify({"success": True, "message": f"Deleted piece: {row['item_name']}"})
    except Exception as e:
        db.rollback()
        return jsonify({"error": str(e)}), 500
    finally:
        cursor.close()
        db.close()


@cashier_bp.after_request
def add_no_cache_headers(response):
    response.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, private"
    response.headers["Pragma"] = "no-cache"
    response.headers["Expires"] = "0"
    return response
